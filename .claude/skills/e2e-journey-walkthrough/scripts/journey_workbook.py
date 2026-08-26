#!/usr/bin/env python3
"""通しテスト用 Excel ワークブックの生成/更新/進捗判定。

blocking: `status` は 全行 PASS(または理由付き SKIP) のときだけ exit 0。1行でも
TODO/FAIL/BLOCKED があれば exit 1 と未達一覧を返す（絶対原則6の実装）。

usage:
  python3 journey_workbook.py init   --out journey.xlsx --plan plan.json
  python3 journey_workbook.py update --out journey.xlsx --id <row_id> \
          --status PASS --evidence <path> --note "..."
  python3 journey_workbook.py status --out journey.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

PLAN_COLS = [
    "id", "order", "role", "phase", "data_condition", "branch", "depends_on",
    "action", "steps", "expected", "status", "evidence", "note",
]
ROLE_COLS = ["role", "how_to_enter", "goal", "provides", "consumes"]
VALID_STATUS = {"TODO", "PASS", "FAIL", "BLOCKED", "SKIP"}
STATUS_FILL = {
    "PASS": "C8E6C9", "FAIL": "FFCDD2", "BLOCKED": "FFE0B2",
    "TODO": "ECEFF1", "SKIP": "E1BEE7",
}
HEAD_FILL = PatternFill("solid", fgColor="263238")
HEAD_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _color_status(ws, status_col: int) -> None:
    for r in range(2, ws.max_row + 1):
        st = str(ws.cell(row=r, column=status_col).value or "TODO")
        fill = STATUS_FILL.get(st)
        if fill:
            ws.cell(row=r, column=status_col).fill = PatternFill("solid", fgColor=fill)


def cmd_init(args) -> int:
    plan = json.loads(Path(args.plan).read_text(encoding="utf-8"))
    rows = sorted(plan.get("rows", []), key=lambda r: (r.get("order", 0), r.get("id", "")))
    wb = Workbook()

    ws = wb.active
    ws.title = "Plan"
    ws.append(PLAN_COLS)
    for r in rows:
        dep = r.get("depends_on", [])
        ws.append([
            r.get("id", ""), r.get("order", 0), r.get("role", ""), r.get("phase", ""),
            r.get("data_condition", ""), r.get("branch", "happy"),
            ",".join(dep) if isinstance(dep, list) else str(dep or ""),
            r.get("action", ""), r.get("steps", ""), r.get("expected", ""),
            r.get("status", "TODO"), r.get("evidence", ""), r.get("note", ""),
        ])
    _style_header(ws, len(PLAN_COLS))
    _color_status(ws, PLAN_COLS.index("status") + 1)
    widths = [22, 6, 12, 12, 14, 14, 20, 30, 34, 34, 10, 26, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).coordinate[0]].width = w

    wr = wb.create_sheet("Roles")
    wr.append(ROLE_COLS)
    discovered = plan.get("discovered", {})
    for role in discovered.get("roles", []):
        meta = (discovered.get("role_meta", {}) or {}).get(role, {})
        wr.append([role, meta.get("how_to_enter", ""), meta.get("goal", ""),
                   meta.get("provides", ""), meta.get("consumes", "")])
    _style_header(wr, len(ROLE_COLS))
    for i, w in enumerate([16, 34, 30, 30, 30], start=1):
        wr.column_dimensions[wr.cell(row=1, column=i).coordinate[0]].width = w

    wb.create_sheet("Summary")
    _write_summary(wb, plan.get("project", ""))
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"init: {len(rows)} rows -> {args.out}")
    return 0


def _write_summary(wb, project: str = "") -> None:
    ws = wb["Plan"]
    sc = PLAN_COLS.index("status") + 1
    counts = {k: 0 for k in ["PASS", "FAIL", "BLOCKED", "TODO", "SKIP"]}
    total = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value in (None, ""):
            continue
        total += 1
        st = str(ws.cell(row=r, column=sc).value or "TODO")
        counts[st] = counts.get(st, 0) + 1
    done = counts["FAIL"] == 0 and counts["BLOCKED"] == 0 and counts["TODO"] == 0 and total > 0
    sm = wb["Summary"]
    sm.delete_rows(1, sm.max_row)
    sm.append(["project", project])
    sm.append(["total", total])
    for k in ["PASS", "FAIL", "BLOCKED", "TODO", "SKIP"]:
        pct = f"{(counts[k] / total * 100):.0f}%" if total else "0%"
        sm.append([k, counts[k], pct])
    sm.append(["DONE?", "YES" if done else "NO"])
    sm.cell(row=sm.max_row, column=2).fill = PatternFill(
        "solid", fgColor="C8E6C9" if done else "FFCDD2")
    sm.cell(row=sm.max_row, column=2).font = Font(bold=True)
    for i, w in enumerate([14, 10, 8], start=1):
        sm.column_dimensions[sm.cell(row=1, column=i).coordinate[0]].width = w


def cmd_update(args) -> int:
    if args.status and args.status not in VALID_STATUS:
        print(f"invalid status: {args.status} (allowed: {sorted(VALID_STATUS)})")
        return 2
    wb = load_workbook(args.out)
    ws = wb["Plan"]
    idx = {c: i + 1 for i, c in enumerate(PLAN_COLS)}
    hit = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=idx["id"]).value) == args.id:
            hit = r
            break
    if hit is None:
        print(f"row id not found: {args.id}")
        return 2
    if args.status:
        ws.cell(row=hit, column=idx["status"]).value = args.status
    if args.evidence is not None:
        ws.cell(row=hit, column=idx["evidence"]).value = args.evidence
    if args.note is not None:
        ws.cell(row=hit, column=idx["note"]).value = args.note
    if args.status == "SKIP" and not (ws.cell(row=hit, column=idx["note"]).value):
        print("SKIP requires a --note reason")
        return 2
    _color_status(ws, idx["status"])
    _write_summary(wb, wb["Summary"].cell(row=1, column=2).value or "")
    wb.save(args.out)
    print(f"update: {args.id} -> {args.status or '(unchanged)'}")
    return 0


def cmd_status(args) -> int:
    wb = load_workbook(args.out)
    ws = wb["Plan"]
    sc = PLAN_COLS.index("status") + 1
    unmet = []
    counts = {}
    total = 0
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=1).value
        if rid in (None, ""):
            continue
        total += 1
        st = str(ws.cell(row=r, column=sc).value or "TODO")
        counts[st] = counts.get(st, 0) + 1
        note = ws.cell(row=r, column=PLAN_COLS.index("note") + 1).value
        if st in ("TODO", "FAIL", "BLOCKED") or (st == "SKIP" and not note):
            unmet.append((rid, st))
    passed = counts.get("PASS", 0)
    print(f"total={total} PASS={passed} "
          f"FAIL={counts.get('FAIL', 0)} BLOCKED={counts.get('BLOCKED', 0)} "
          f"TODO={counts.get('TODO', 0)} SKIP={counts.get('SKIP', 0)} "
          f"({(passed / total * 100):.0f}% pass)" if total else "empty plan")
    if unmet:
        print(f"NOT DONE — {len(unmet)} row(s) unmet:")
        for rid, st in unmet[:50]:
            print(f"  [{st}] {rid}")
        return 1
    print("ALL PASS — journey complete.")
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description="通しテスト Excel ワークブック")
    sub = p.add_subparsers(dest="cmd", required=True)
    pi = sub.add_parser("init"); pi.add_argument("--out", required=True); pi.add_argument("--plan", required=True)
    pu = sub.add_parser("update")
    pu.add_argument("--out", required=True); pu.add_argument("--id", required=True)
    pu.add_argument("--status"); pu.add_argument("--evidence"); pu.add_argument("--note")
    psx = sub.add_parser("status"); psx.add_argument("--out", required=True)
    args = p.parse_args()
    return {"init": cmd_init, "update": cmd_update, "status": cmd_status}[args.cmd](args)


if __name__ == "__main__":
    sys.exit(main())
