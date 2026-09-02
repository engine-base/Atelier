#!/usr/bin/env python3
"""build_xlsx.py — テスト仕様書を「2系統の Excel」で決定論生成する（鉄則3+鉄則4の実体）。

- 入力: 画面別 spec（`<spec-dir>/screens/*.md`。各 md 先頭 `# <ID> <日本語画面名> 画面別...`、
  本文に 9 列表 `| ID | 画面 | テスト観点 | テスト項目 | 前提条件 | 操作手順 | 期待結果 | 結果 | 備考 |`、
  または test-ladder (L1〜L5) 対応の 11 列表 `… | 備考 | タスク | 実行条件 |`。9 列の旧行も読める）。
- 出力（必ず 2 冊）:
  1. `<out>/テスト仕様書_クライアント版.xlsx` … 画面で操作・目視できる TC のみ・平易日本語・日本語画面名タブ・結果=完了/空。
  2. `<out>/テスト仕様書_エンジニア版.xlsx` … 全 TC・原文・結果ステータス色分け・概要タブ・
     **実行計画タブ**（段 L1〜L5 × 担当タスク × 実行条件。test-ladder.md §3）。
- クライアント版は「画面で見えない技術専用 TC」を除外し、除外件数を概要タブに明示（完全性の証明）。

使い方:
  python3 build_xlsx.py <spec-dir> [out-dir]
    <spec-dir> 例: .qa/test-specs        (screens/ を含むフォルダ)
    out-dir 省略時は <spec-dir>/../ (=.qa) に出力。

依存: openpyxl（無ければ `pip3 install --break-system-packages openpyxl`）。
"""
import re
import sys
import os
import glob

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl が必要です: pip3 install --break-system-packages openpyxl")

TECH_KANTEN = ['権限', '死にボタン', '死にリンク', 'クロス', '契約', '環境変数', 'セキュリティ', 'RLS']
TECH_TXT = re.compile(
    r'(DB突合|select |curl|コードを読|環境変数|鍵|マイグレーション|契約テスト|単体テスト|'
    r'ルート実在|service_role|network|console|Playwright)')
# クライアント版の平易化（裏側語→平易語。references/client-engineer-views.md の言い換え表に対応）
REPL = [
    (r'死にボタン検査[:：]?', 'ボタンが正しく動作するか'),
    (r'死にリンク[:：]?', '画面の導線が切れていないか'),
    (r'/api/v1/[^\s、。]*', ''), (r'/api/[^\s、。]*', ''),
    (r'HTTP\s?\d{3}', ''), (r'\b(200|201|204|400|401|403|422|429|500)\b', ''),
    (r'role=alert', 'エラー表示'),
    (r'\bRLS\b|行レベル権限', '他人のデータが見えない権限'),
    (r'レスポンシブ', 'スマホ・タブレットでも崩れない'),
    (r'\ba11y\b|アクセシビリティ', '読み上げ対応'),
    (r'\b(POST|GET|DELETE|PATCH|PUT)\b', ''),
    (r'network|console|Playwright|DB突合|DBクエリ', ''),
    (r'router\.(push|replace)|遷移する', '次の画面に進む'),
    (r'冪等|idempotent', '繰り返しても安全'),
]


def is_tech(kanten, koumoku, exp, biko):
    if any(t in kanten for t in TECH_KANTEN):
        return True
    return bool(TECH_TXT.search(koumoku + exp + biko))


def plain(s):
    for pat, rep in REPL:
        s = re.sub(pat, rep, s)
    return re.sub(r'\s{2,}', ' ', s).strip(' 、')


def sheetname(nm):
    return re.sub(r'[\\/*?:\[\]]', '', nm)[:31] or 'sheet'


def done(res):
    return 'PASS' in res or 'FIXED' in res or 'passed' in res or '完了' in res


HDR_FILL = PatternFill('solid', fgColor='7A5C9E')
HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
DONE_FILL = PatternFill('solid', fgColor='E8F5E9')
BLK_FILL = PatternFill('solid', fgColor='FFF3E0')
THIN = Side(style='thin', color='DDDDDD')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDER
            c.font = Font(size=9)


def load(spec_dir):
    data = []
    for fp in sorted(glob.glob(os.path.join(spec_dir, 'screens', '*.md'))):
        lines = open(fp, encoding='utf-8').read().split('\n')
        m = re.match(r'#\s*([A-Za-z]+\d*)\s+(.+?)\s*画面別', lines[0]) if lines else None
        sid = m.group(1) if m else os.path.basename(fp)[:-3]
        jp = m.group(2) if m else sid
        rows = []
        for ln in lines:
            if not re.match(r'^\|\s*[A-Za-z]+\d*-\d{3}\s*\|', ln):
                continue
            c = [x.strip() for x in ln.split('|')[1:-1]]
            if len(c) >= 9:
                # 11 列 (タスク / 実行条件) は test-ladder。9 列の旧行は空で埋める
                rows.append((c[:11] + ['', ''])[:11])
        if rows:
            data.append((jp, sid, rows))
    return data


def build_client(data, path):
    wb = Workbook()
    wb.remove(wb.active)
    ov = wb.create_sheet('概要')
    ov.append(['テスト仕様書（クライアント確認用）'])
    ov.append(['画面で操作・目視できる項目のみ・専門用語なし。「結果」欄は確認後に○を付けてご利用ください。'])
    ov.append([])
    ov.append(['画面', '確認項目数', 'うち動作確認済(完了)'])
    tot = don = allc = 0
    for jp, sid, rows in data:
        allc += len(rows)
        vis = [r for r in rows if not is_tech(r[2], r[3], r[6], r[8])]
        d = sum(1 for r in vis if done(r[7]))
        tot += len(vis)
        don += d
        ov.append([f'{jp}画面', len(vis), d])
        ws = wb.create_sheet(sheetname(jp))
        ws.append(['ID', '画面', 'テスト観点', 'テスト項目', '前提条件', '操作手順', '期待結果', '結果', '備考'])
        for r in vis:
            ws.append([r[0], f'{jp}画面', plain(r[2]), plain(r[3]), plain(r[4]),
                       plain(r[5]), plain(r[6]), '完了' if done(r[7]) else '', ''])
        style(ws, [10, 12, 12, 26, 20, 26, 30, 7, 16])
        for row in ws.iter_rows(min_row=2):
            if row[7].value == '完了':
                for c in row:
                    c.fill = DONE_FILL
    ov.append([])
    ov.append([f'完全性: 全{allc}件 ＝ クライアント可視{tot}件 ＋ 技術専用{allc-tot}件（画面で見えない裏取り=エンジニア版）'])
    style(ov, [24, 12, 18])
    ov['A1'].font = Font(bold=True, size=13, color='4E3A67')
    wb.save(path)
    return tot, don, allc


def build_eng(data, path):
    wb = Workbook()
    wb.remove(wb.active)
    ov = wb.create_sheet('概要')
    ov.append(['テスト仕様書（エンジニア版・全件）'])
    ov.append(['全TC＋結果ステータス。証拠は runs/*/findings.md 参照。'])
    ov.append([])
    ov.append(['画面', '総数', 'PASS', 'BLOCKED', '要実走'])
    for jp, sid, rows in data:
        p = sum(1 for r in rows if done(r[7]))
        b = sum(1 for r in rows if 'BLOCKED' in r[7])
        ov.append([f'{jp}({sid})', len(rows), p, b, len(rows) - p - b])
        ws = wb.create_sheet(sheetname(f'{sid}_{jp}'))
        ws.append(['ID', '画面', 'テスト観点', 'テスト項目', '前提条件', '操作手順', '期待結果', '結果', '備考',
                   'タスク', '実行条件'])
        for r in rows:
            ws.append(r)
        style(ws, [10, 10, 12, 26, 18, 24, 28, 16, 16, 12, 14])
        for row in ws.iter_rows(min_row=2):
            v = str(row[7].value)
            if done(v):
                for c in row:
                    c.fill = DONE_FILL
            elif 'BLOCKED' in v:
                for c in row:
                    c.fill = BLK_FILL
    style(ov, [22, 8, 8, 10, 10])
    ov['A1'].font = Font(bold=True, size=13, color='4E3A67')
    build_plan_sheet(wb, data)
    wb.save(path)


def level_of(cond):
    m = re.search(r'\bL([1-5])\b', cond or '')
    return int(m.group(1)) if m else 1


def build_plan_sheet(wb, data):
    """実行計画タブ: 段 (L1〜L5) × 担当タスク × 実行条件 で全行を並べ直す (test-ladder.md §3)。

    画面タブは「1 画面が正しいか」を見る形。ここは「誰が・いつ流すか」を見る形。
    タスク列が空の L1 行は『未割当』として先頭に集め、後追い適用の残りが一目で分かるようにする。
    """
    ws = wb.create_sheet('実行計画', 1)
    ws.append(['段', 'タスク', '実行条件', 'ID', '画面', 'テスト項目', '結果', '備考'])
    flat = []
    for jp, sid, rows in data:
        for r in rows:
            task = r[9] or '（未割当）'
            cond = r[10] or 'L1'
            flat.append((level_of(cond), task == '（未割当）' and ' ' or task, cond, r[0], jp, r[3], r[7], r[8]))
    flat.sort(key=lambda x: (x[0], x[1], x[3]))
    for lv, task, cond, rid, jp, item, res, biko in flat:
        ws.append([f'L{lv}', task.strip() or '（未割当）', cond, rid, jp, item, res, biko])
    style(ws, [5, 12, 16, 12, 12, 30, 16, 16])
    for row in ws.iter_rows(min_row=2):
        v = str(row[6].value)
        if done(v):
            for c in row:
                c.fill = DONE_FILL
        elif 'BLOCKED' in v:
            for c in row:
                c.fill = BLK_FILL
    # 段別サマリを概要タブに足す
    ov = wb['概要']
    ov.append([])
    ov.append(['段', '行数', 'PASS', 'タスク未割当'])
    for lv in range(1, 6):
        rs = [x for x in flat if x[0] == lv]
        if rs:
            ov.append([f'L{lv}', len(rs), sum(1 for x in rs if done(x[6])), sum(1 for x in rs if not x[1].strip())])


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    spec_dir = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(spec_dir.rstrip('/'))
    data = load(spec_dir)
    if not data:
        sys.exit(f'screens/*.md が見つかりません: {spec_dir}/screens/')
    os.makedirs(out, exist_ok=True)
    cpath = os.path.join(out, 'テスト仕様書_クライアント版.xlsx')
    epath = os.path.join(out, 'テスト仕様書_エンジニア版.xlsx')
    tot, don, allc = build_client(data, cpath)
    build_eng(data, epath)
    print(f'✓ クライアント版: {cpath} ({tot}件・完了{don})')
    print(f'✓ エンジニア版 : {epath} (全{allc}件)')
    print(f'完全性: 全{allc} ＝ クライアント可視{tot} ＋ 技術専用{allc-tot}')


if __name__ == '__main__':
    main()
