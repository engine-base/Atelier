"""GAP-163: Excel / CSV 成果物のツール内表示と編集。

経営者質問 (2026-08-19):
  「あとエクセルとかだとここの表示はどうなるの？？
   エクセルとかスプレッドシートをここで修正とかできるの？PDF もだけど」

方針 (正直な線引き):
  - **Excel / CSV** … シートを表として表示し、**セル編集して新バージョン保存**できる。
    保存は元の形式 (xlsx) で書き戻すので、そのままクライアントに渡せる。
    数式・書式・グラフは保持しない (値のみ) — その旨を画面と API で明示する。
  - **PDF** … ブラウザ内蔵ビューアで**表示**する (既に inline 配信済)。
    直接の編集はしない。修正は「元の成果物 (HTML) を AI に直してもらって出し直す」。
    PDF を直接書き換える機能は**あるふりをしない**。
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from src.audit import AuditEvent, AuditWriter

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MIME = "text/csv"

MAX_ROWS = 500
MAX_COLS = 60


class SheetError(Exception):
    """code: not_found / unsupported / too_large"""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


@dataclass(frozen=True)
class SheetData:
    file_name: str
    mime: str
    editable: bool
    #: [{"name": str, "rows": [[str, ...], ...]}]
    sheets: list[dict[str, object]]
    note: str = ""
    #: GAP-254: この表の版 (保存時に base_version として返してもらう)
    version: int = 1


def _rows_from_xlsx(data: bytes) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    out: list[dict[str, object]] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS:
                break
            rows.append(["" if c is None else str(c) for c in row[:MAX_COLS]])
        out.append({"name": ws.title, "rows": rows})
    wb.close()
    return out


def _rows_from_csv(data: bytes) -> list[dict[str, object]]:
    for enc in ("utf-8-sig", "cp932", "utf-16"):
        try:
            text_body = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text_body = data.decode("utf-8", errors="replace")
    rows = [r[:MAX_COLS] for i, r in enumerate(csv.reader(io.StringIO(text_body))) if i < MAX_ROWS]
    return [{"name": "CSV", "rows": rows}]


def rows_to_xlsx(sheets: list[dict[str, object]]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)  # pyright: ignore[reportArgumentType]
    for i, sheet in enumerate(sheets, start=1):
        name = str(sheet.get("name") or f"Sheet{i}")[:28] or f"Sheet{i}"
        ws = wb.create_sheet(title=name)
        raw_rows = sheet.get("rows")
        rows = raw_rows if isinstance(raw_rows, list) else []
        for row in rows[:MAX_ROWS]:
            cells = row if isinstance(row, list) else []
            ws.append([str(c) for c in cells[:MAX_COLS]])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def load_sheet(session: AsyncSession, *, output_id: str) -> SheetData:
    """成果物を表として読む。Excel/CSV 以外は unsupported (推測で表にしない)。"""
    from src.services.mocks.artifacts import FILEDB_PREFIX, fetch_file_service

    row = (
        await session.execute(
            text(
                "select html_path, version from public.workflow_outputs "
                "where id = cast(:i as uuid) and deleted_at is null"
            ),
            {"i": output_id},
        )
    ).first()
    if row is None or row.html_path is None:
        raise SheetError("not_found", "成果物が見つかりません")
    path = str(row.html_path)
    if not path.startswith(FILEDB_PREFIX):
        raise SheetError(
            "unsupported",
            "この成果物は表形式ではありません (HTML 成果物は本文プレビューで開けます)",
        )
    found = await fetch_file_service(path[len(FILEDB_PREFIX) :])
    if found is None:
        raise SheetError("not_found", "ファイルの実体が見つかりません")
    data, mime, file_name = found
    lower = file_name.lower()
    if mime == XLSX_MIME or lower.endswith((".xlsx", ".xlsm")):
        return SheetData(
            version=int(row.version),
            file_name=file_name,
            mime=XLSX_MIME,
            editable=True,
            sheets=_rows_from_xlsx(data),
            note="値のみを表示・編集します (数式・書式・グラフは保持されません)",
        )
    if mime == CSV_MIME or lower.endswith(".csv"):
        return SheetData(
            version=int(row.version),
            file_name=file_name,
            mime=CSV_MIME,
            editable=True,
            sheets=_rows_from_csv(data),
            note="CSV を表として表示・編集します",
        )
    if mime == "application/pdf" or lower.endswith(".pdf"):
        # GAP-225: 「表として扱えない」一般形と分けて持つ。route は文言を
        # code から引くので、同じ code のままだと **PDF だけの案内 (見られるが
        # 直せない・どう直すか) が消える**。
        raise SheetError("pdf_view_only", "PDF は表として編集できない")
    raise SheetError("unsupported", f"この形式は表として扱えません ({mime})")


async def save_sheet(
    session: AsyncSession,
    *,
    actor_id: str,
    output_id: str,
    sheets: list[dict[str, object]],
    base_version: int | None = None,
) -> str | None:
    """編集内容を **新バージョン** として保存する (元の版は残す)。返り値 = 新 output id。

    GAP-254: base_version (編集を始めた時点の版) がチェーンの最新と違えば version_conflict (→ 409)。
    2 つのタブで同じ成果物を編集して両方保存すると、後から保存した古い内容が黙って最新版になっていた。
    """
    from src.services.mocks.artifacts import FILEDB_PREFIX, store_file_service
    from src.services.outputs import get_output, insert_version, list_versions

    current = await get_output(session, output_id)
    if current is None:
        return None
    if base_version is not None:
        latest = max(
            (v.version for v in await list_versions(session, output_id)), default=current.version
        )
        if base_version != latest:
            raise SheetError(
                "version_conflict",
                f"base_version={base_version} but latest is v{latest}",
            )
    data = await load_sheet(session, output_id=output_id)
    if not data.editable:
        raise SheetError("not_editable", "この成果物は編集できない形式")
    file_id = await store_file_service(
        data=rows_to_xlsx(sheets),
        mime=XLSX_MIME,
        file_name=(
            data.file_name if data.file_name.lower().endswith(".xlsx") else f"{data.file_name}.xlsx"
        ),
    )
    created = await insert_version(
        session,
        src=current,
        html_path=f"{FILEDB_PREFIX}{file_id}",
        meta={
            "author": "user",
            "edit": "sheet",
            "file_name": data.file_name,
            "note": "表の値を編集",
        },
        actor_id=actor_id,
    )
    await AuditWriter(session).write(
        AuditEvent(
            action="output.sheet.edit",
            target_type="workflow_output",
            actor_type="user",
            actor_id=actor_id,
            target_id=created.id,
            after={"source_output_id": output_id, "version": created.version},
        )
    )
    return created.id
