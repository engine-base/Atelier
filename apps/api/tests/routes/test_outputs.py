"""Integration tests for /outputs (T-A-21) — 実 Postgres + RLS + JWT。実 DB 無なら skip。"""

from __future__ import annotations

import asyncio
import base64
import hashlib
import hmac
import json
import os
import time
import uuid
from collections.abc import AsyncGenerator, Iterator
from typing import Annotated

import pytest

PG_ASYNC = os.environ.get(
    "ATELIER_TEST_PG_URL", "postgresql+asyncpg://postgres@/postgres?host=/tmp&port=54322"
)
PG_SYNC = PG_ASYNC.replace("+asyncpg", "+psycopg")
JWT_SECRET = "test-jwt-secret"
os.environ.setdefault("ATELIER_AUTH_JWT_SECRET", JWT_SECRET)

import sqlalchemy  # noqa: E402
from fastapi import Depends, FastAPI  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from sqlalchemy import text  # noqa: E402
from sqlalchemy.ext.asyncio import (  # noqa: E402
    AsyncEngine,
    AsyncSession,
    create_async_engine,
)
from sqlalchemy.pool import NullPool  # noqa: E402

from src.dependencies import CurrentUser, get_current_user, get_rls_session  # noqa: E402
from tests.routes._fixtures import ensure_ai_employee  # noqa: E402


def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode("ascii")


def _mint_jwt(user_id: str) -> str:
    header = _b64url(json.dumps({"alg": "HS256", "typ": "JWT"}).encode())
    payload = _b64url(
        json.dumps(
            {
                "sub": user_id,
                "role": "authenticated",
                "aud": "authenticated",
                "exp": int(time.time()) + 3600,
            }
        ).encode()
    )
    sig = _b64url(
        hmac.new(
            JWT_SECRET.encode(), f"{header}.{payload}".encode("ascii"), hashlib.sha256
        ).digest()
    )
    return f"{header}.{payload}.{sig}"


def _db_available() -> bool:
    try:
        eng = sqlalchemy.create_engine(PG_SYNC, poolclass=NullPool)
        try:
            with eng.connect() as c:
                c.execute(text("select 1"))
        finally:
            eng.dispose()
        return True
    except Exception:
        return False


pytestmark = pytest.mark.skipif(not _db_available(), reason="local Postgres not available")


@pytest.fixture()
def app() -> Iterator[FastAPI]:
    test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)

    async def _override_session(
        user: Annotated[CurrentUser, Depends(get_current_user)],
    ) -> AsyncGenerator[AsyncSession, None]:
        claims = json.dumps({"sub": user.id, "role": user.role})
        async with AsyncSession(test_engine) as session:
            await session.execute(
                text("select set_config('request.jwt.claims', :c, true)"), {"c": claims}
            )
            await session.execute(text("set local role authenticated"))
            try:
                yield session
            except Exception:
                await session.rollback()
                raise
            else:
                await session.commit()

    from src.routes import api_router

    application = FastAPI()
    application.include_router(api_router)
    application.dependency_overrides[get_rls_session] = _override_session
    yield application
    asyncio.run(test_engine.dispose())


@pytest.fixture()
def sync_engine() -> Iterator[sqlalchemy.Engine]:
    eng = sqlalchemy.create_engine(PG_SYNC, poolclass=NullPool)
    yield eng
    eng.dispose()


@pytest.fixture()
def seeded(sync_engine: sqlalchemy.Engine) -> Iterator[dict[str, str]]:
    u_a, u_b = str(uuid.uuid4()), str(uuid.uuid4())
    ws_a, ws_b = str(uuid.uuid4()), str(uuid.uuid4())
    proj_a, out_a = str(uuid.uuid4()), str(uuid.uuid4())
    with sync_engine.begin() as c:
        for uid in (u_a, u_b):
            em = f"ta21-{uid[:8]}@t.invalid"
            c.execute(text("insert into auth.users (id,email) values (:i,:e)"), {"i": uid, "e": em})
            c.execute(
                text("insert into public.users (id,email) values (:i,:e)"), {"i": uid, "e": em}
            )
        for ws, owner in ((ws_a, u_a), (ws_b, u_b)):
            c.execute(
                text("insert into public.workspaces (id,owner_user_id,name) values (:i,:o,:n)"),
                {"i": ws, "o": owner, "n": f"ws-{ws[:6]}"},
            )
        c.execute(
            text(
                "insert into public.projects (id,workspace_id,name,project_type) values (:i,:w,:n,'internal_product')"
            ),
            {"i": proj_a, "w": ws_a, "n": "proj-a"},
        )
        # 成果物は工程生成で作られるため、service_role 相当 (superuser) で seed
        c.execute(
            text(
                "insert into public.workflow_outputs (id,project_id,stage,summary) values (cast(:i as uuid),cast(:p as uuid),'design','sum')"
            ),
            {"i": out_a, "p": proj_a},
        )
    yield {"u_a": u_a, "u_b": u_b, "ws_a": ws_a, "proj_a": proj_a, "out_a": out_a}
    with sync_engine.begin() as c:
        c.execute(text("delete from public.workspaces where id in (:a,:b)"), {"a": ws_a, "b": ws_b})
        c.execute(text("delete from public.users where id in (:a,:b)"), {"a": u_a, "b": u_b})
        c.execute(text("delete from auth.users where id in (:a,:b)"), {"a": u_a, "b": u_b})


def _h(uid: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {_mint_jwt(uid)}"}


_ANCHOR_HTML = (
    "<html><body>"
    '<h2 id="sec-1">1. プロジェクト概要</h2><p>本文</p>'
    '<h2 id="sec-2">2. 成功の定義</h2>'
    "</body></html>"
)


def _install_storage_fakes(
    monkeypatch: pytest.MonkeyPatch, *, source_html: str = _ANCHOR_HTML
) -> list[str]:
    """outputs revise 経路の storage HTTP 境界をフェイク化する (GAP-023 テスト用)。

    署名 (post) / 取得 (get) / アップロード (put) を 1 つのフェイクで担い、
    アップロードされた HTML 本文の記録リストを返す。
    """
    from typing import Any

    from src import storage_signing
    from src.services.outputs import revise as revise_svc

    monkeypatch.setenv("ATELIER_SUPABASE_ADMIN_API_URL", "https://stor.invalid")
    monkeypatch.setenv("ATELIER_SUPABASE_SERVICE_ROLE_KEY", "svc-key")
    uploaded: list[str] = []

    class _Res:
        def __init__(self, payload: dict[str, Any] | None = None, body_text: str = "") -> None:
            self.status_code = 200
            self._payload = payload or {}
            self.text = body_text

        def json(self) -> dict[str, Any]:
            return self._payload

    class _StorageClient:
        def __init__(self, *_a: Any, **_k: Any) -> None: ...

        async def __aenter__(self) -> _StorageClient:
            return self

        async def __aexit__(self, *_a: Any) -> bool:
            return False

        async def post(self, url: str, **_k: Any) -> _Res:
            if "/object/upload/sign/" in url:
                return _Res({"url": "/object/upload/sign/outputs/x?token=u"})
            return _Res({"signedURL": "/object/sign/outputs/x?token=d"})

        async def get(self, _url: str, **_k: Any) -> _Res:
            return _Res(body_text=source_html)

        async def put(self, _url: str, *, content: bytes, **_k: Any) -> _Res:
            uploaded.append(content.decode("utf-8"))
            return _Res()

    monkeypatch.setattr(storage_signing.httpx, "AsyncClient", _StorageClient)
    assert revise_svc.httpx.AsyncClient is _StorageClient  # 共有モジュールの前提を明示
    return uploaded


@pytest.mark.integration
class TestOutputs:
    def test_unauthenticated_401(self, app: FastAPI) -> None:
        with TestClient(app) as client:
            assert client.get("/outputs").status_code == 401

    def test_list_and_get(self, app: FastAPI, seeded: dict[str, str]) -> None:
        h = _h(seeded["u_a"])
        with TestClient(app) as client:
            lst = client.get(f"/outputs?project_id={seeded['proj_a']}", headers=h)
            assert lst.status_code == 200
            assert any(x["id"] == seeded["out_a"] for x in lst.json()["data"])
            # stage filter
            assert any(
                x["id"] == seeded["out_a"]
                for x in client.get(
                    f"/outputs?project_id={seeded['proj_a']}&stage=design", headers=h
                ).json()["data"]
            )
            g = client.get(f"/outputs/{seeded['out_a']}", headers=h)
            assert g.status_code == 200
            assert g.json()["data"]["stage"] == "design"

    def test_cross_workspace_invisible_404(self, app: FastAPI, seeded: dict[str, str]) -> None:
        hb = _h(seeded["u_b"])
        with TestClient(app) as client:
            assert client.get(f"/outputs/{seeded['out_a']}", headers=hb).status_code == 404
            assert all(
                x["id"] != seeded["out_a"]
                for x in client.get(f"/outputs?project_id={seeded['proj_a']}", headers=hb).json()[
                    "data"
                ]
            )


def _seed_output(
    sync_engine: sqlalchemy.Engine,
    seeded: dict[str, str],
    *,
    html_path: str | None = "outputs/g01test/doc-v1.html",
    md_path: str | None = None,
    version: int = 1,
) -> str:
    oid = str(uuid.uuid4())
    with sync_engine.begin() as c:
        c.execute(
            text(
                "insert into public.workflow_outputs "
                "(id,project_id,stage,summary,html_path,md_path,version) "
                "values (cast(:i as uuid),cast(:p as uuid),'requirements','要件定義書',"
                ":hp,:mp,:v)"
            ),
            {"i": oid, "p": seeded["proj_a"], "hp": html_path, "mp": md_path, "v": version},
        )
    return oid


@pytest.mark.integration
class TestOutputViewerOps:
    """GAP-023: format 別 content-url / versions / anchors / revise / fix-proposals。"""

    def test_content_url_format_md_and_missing_409(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        _install_storage_fakes(monkeypatch)
        oid = _seed_output(sync_engine, seeded, md_path="outputs/g01test/doc-v1.md")
        h = _h(seeded["u_a"])
        with TestClient(app) as client:
            assert client.get(f"/outputs/{oid}/content-url?format=md", headers=h).status_code == 200
            assert client.get(f"/outputs/{oid}/content-url", headers=h).status_code == 200
            # json は未生成 → 409 (存在しない版を偽装しない)
            r = client.get(f"/outputs/{oid}/content-url?format=json", headers=h)
            assert r.status_code == 409

    def test_versions_chain_and_cross_ws_404(
        self, app: FastAPI, seeded: dict[str, str], sync_engine: sqlalchemy.Engine
    ) -> None:
        o1 = _seed_output(sync_engine, seeded, version=1)
        o2 = _seed_output(sync_engine, seeded, version=2)
        with TestClient(app) as client:
            r = client.get(f"/outputs/{o1}/versions", headers=_h(seeded["u_a"]))
            assert r.status_code == 200
            versions = [x["version"] for x in r.json()["data"] if x["id"] in (o1, o2)]
            assert versions == [1, 2]
            assert (
                client.get(f"/outputs/{o1}/versions", headers=_h(seeded["u_b"])).status_code == 404
            )

    def test_anchors_extracted_from_real_html(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        _install_storage_fakes(monkeypatch)
        oid = _seed_output(sync_engine, seeded)
        with TestClient(app) as client:
            r = client.get(f"/outputs/{oid}/anchors", headers=_h(seeded["u_a"]))
            assert r.status_code == 200
            anchors = {a["element_id"]: a["label"] for a in r.json()["data"]}
            assert anchors["sec-1"] == "1. プロジェクト概要"
            assert "sec-2" in anchors
            # HTML 未生成は 409
            no_html = _seed_output(sync_engine, seeded, html_path=None, version=9)
            assert (
                client.get(f"/outputs/{no_html}/anchors", headers=_h(seeded["u_a"])).status_code
                == 409
            )

    def test_revise_fake_llm_creates_steve_version(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        uploaded = _install_storage_fakes(monkeypatch)
        monkeypatch.setenv("ATELIER_ALLOW_FAKE_LLM", "1")
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        oid = _seed_output(sync_engine, seeded, md_path="outputs/g01test/doc-v1.md")
        h = _h(seeded["u_a"])
        with TestClient(app) as client:
            r = client.post(
                f"/outputs/{oid}/revise",
                json={"instruction": "2.5 項に可視範囲サブセクションを追加"},
                headers=h,
            )
            assert r.status_code == 201, r.text
            v2 = r.json()["data"]
            assert v2["version"] == 2
            assert v2["meta"]["author"] == "steve"
            assert v2["meta"]["revision_instruction"] == "2.5 項に可視範囲サブセクションを追加"
            # GAP-171: meta には「どの実行経路 = 誰の費用で作られたか」を記録する
            # (relay = 本人の Claude サブスク / agent_sdk / api / fake)。
            assert v2["meta"]["model"] == "fake"
            assert v2["html_path"].endswith("-rev.html")
            # 改訂は HTML に対して行われるため json/md は未生成 (旧版を偽装しない)
            assert v2["json_path"] is None and v2["md_path"] is None
            assert len(uploaded) == 1 and 'data-fake-revision="1"' in uploaded[0]
            with sync_engine.connect() as c:
                n = c.execute(
                    text(
                        "select count(*) from public.audit_logs "
                        "where action='output.revise' and target_id=:t"
                    ),
                    {"t": v2["id"]},
                ).scalar_one()
            assert n == 1

    def test_revise_503_and_409_no_html(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        _install_storage_fakes(monkeypatch)
        monkeypatch.delenv("ATELIER_ALLOW_FAKE_LLM", raising=False)
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        oid = _seed_output(sync_engine, seeded)
        no_html = _seed_output(sync_engine, seeded, html_path=None, version=8)
        h = _h(seeded["u_a"])
        with TestClient(app) as client:
            assert (
                client.post(f"/outputs/{oid}/revise", json={"instruction": "x"}, headers=h)
            ).status_code == 503
            assert (
                client.post(f"/outputs/{no_html}/revise", json={"instruction": "x"}, headers=h)
            ).status_code == 409
            # R-T08: 他 WS からは 404
            monkeypatch.setenv("ATELIER_ALLOW_FAKE_LLM", "1")
            assert (
                client.post(
                    f"/outputs/{oid}/revise",
                    json={"instruction": "越境"},
                    headers=_h(seeded["u_b"]),
                )
            ).status_code == 404
            # 不正 UUID は 500 ではなく 404 (実監査が検出した 500 の回帰ガード)
            assert (
                client.post("/outputs/not-a-uuid/revise", json={"instruction": "x"}, headers=h)
            ).status_code == 404
            assert (
                client.post("/comments/optimistic-2/fix-proposal", headers=h)
            ).status_code == 404
            assert (client.post("/output-fix-proposals/junk/approve", headers=h)).status_code == 404

    def test_fix_proposal_lifecycle(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """提案生成 (pending, 重複 409) → 承認 (新バージョン + approved) / 却下 (不変)。"""
        _install_storage_fakes(monkeypatch)
        monkeypatch.setenv("ATELIER_ALLOW_FAKE_LLM", "1")
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        oid = _seed_output(sync_engine, seeded)
        h = _h(seeded["u_a"])
        with TestClient(app) as client:
            c1 = client.post(
                "/comments",
                json={
                    "target_type": "workflow_output",
                    "target_id": oid,
                    "content": "可視範囲を明示してほしい",
                    "target_element_id": "sec-1",
                },
                headers=h,
            ).json()["data"]
            assert c1["target_element_id"] == "sec-1"

            p = client.post(f"/comments/{c1['id']}/fix-proposal", headers=h)
            assert p.status_code == 201, p.text
            prop = p.json()["data"]
            assert prop["status"] == "pending"
            assert "可視範囲を明示してほしい" in prop["proposal"]
            # 同一コメントへの pending 重複は 409
            assert client.post(f"/comments/{c1['id']}/fix-proposal", headers=h).status_code == 409

            a = client.post(f"/output-fix-proposals/{prop['id']}/approve", headers=h)
            assert a.status_code == 200, a.text
            body = a.json()["data"]
            assert body["proposal"]["status"] == "approved"
            assert body["new_output"]["version"] == 2
            assert body["proposal"]["applied_output_id"] == body["new_output"]["id"]
            # 二重承認は 409。GAP-260: 理由は「処理済み」であって同時編集の文言ではない
            dup = client.post(f"/output-fix-proposals/{prop['id']}/approve", headers=h)
            assert dup.status_code == 409
            assert "処理済み" in dup.json()["detail"], dup.text
            assert "同時に保存" not in dup.json()["detail"], dup.text

            # 却下: 文書は不変 (バージョン数が増えない)
            c2 = client.post(
                "/comments",
                json={"target_type": "workflow_output", "target_id": oid, "content": "別件"},
                headers=h,
            ).json()["data"]
            p2 = client.post(f"/comments/{c2['id']}/fix-proposal", headers=h).json()["data"]
            before = len(client.get(f"/outputs/{oid}/versions", headers=h).json()["data"])
            rj = client.post(f"/output-fix-proposals/{p2['id']}/reject", headers=h)
            assert rj.status_code == 200
            assert rj.json()["data"]["status"] == "rejected"
            after = len(client.get(f"/outputs/{oid}/versions", headers=h).json()["data"])
            assert before == after

            lst = client.get(f"/outputs/{oid}/fix-proposals", headers=h)
            assert lst.status_code == 200
            statuses = sorted(x["status"] for x in lst.json()["data"])
            assert statuses == ["approved", "rejected"]
            # R-T08: 他 WS からは提案一覧も 404
            assert (
                client.get(f"/outputs/{oid}/fix-proposals", headers=_h(seeded["u_b"])).status_code
                == 404
            )
            with sync_engine.connect() as c:
                n = c.execute(
                    text(
                        "select count(*) from public.audit_logs where action in "
                        "('output.fix_proposal.propose','output.fix_proposal.approve',"
                        "'output.fix_proposal.reject') and target_id in (:o, :n2)"
                    ),
                    {"o": oid, "n2": body["new_output"]["id"]},
                ).scalar_one()
            assert n >= 3


# ── GAP-155: 成果物の差分 + 復元 + 同時改訂ガード ──────────────────────────

_OLD_HTML = "<html>\n<body>\n<h2>1. 概要</h2>\n<p>旧仕様</p>\n</body>\n</html>"
_NEW_HTML = "<html>\n<body>\n<h2>1. 概要</h2>\n<p>新仕様</p>\n<p>追記</p>\n</body>\n</html>"


def _seed_output_version(
    sync_engine: sqlalchemy.Engine,
    *,
    project_id: str,
    stage: str,
    html: str | None,
    version: int,
    path_override: str | None = None,
    meta: str = "{}",
) -> str:
    """mock_contents に実 HTML を置き mockdb:// 参照の workflow_outputs 行を作る。"""
    oid = str(uuid.uuid4())
    with sync_engine.begin() as c:
        path = path_override
        if html is not None:
            cid = c.execute(
                text("insert into public.mock_contents (html) values (:h) returning id"),
                {"h": html},
            ).scalar_one()
            path = f"mockdb://{cid}"
        c.execute(
            text(
                "insert into public.workflow_outputs "
                "(id, project_id, stage, html_path, summary, version, meta) "
                "values (cast(:i as uuid), cast(:p as uuid), cast(:s as workflow_stage_enum), "
                "        :path, 'sum', :v, cast(:m as jsonb))"
            ),
            {"i": oid, "p": project_id, "s": stage, "path": path, "v": version, "m": meta},
        )
    return oid


@pytest.mark.integration
class TestGap155OutputsDiffRestore:
    def _patch_service_factory(self, monkeypatch: pytest.MonkeyPatch) -> AsyncEngine:
        from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

        from src.services.mocks import artifacts as artifacts_svc

        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        monkeypatch.setattr(
            artifacts_svc,
            "service_session_factory",
            lambda: async_sessionmaker(test_engine, class_=AsyncSession),
        )
        return test_engine

    def test_diff_restore_roundtrip(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """差分は実 HTML から計算し、復元は履歴を消さず新版として積む。"""
        test_engine = self._patch_service_factory(monkeypatch)
        proj = seeded["proj_a"]
        v1 = _seed_output_version(
            sync_engine, project_id=proj, stage="requirements", html=_OLD_HTML, version=1
        )
        v2 = _seed_output_version(
            sync_engine, project_id=proj, stage="requirements", html=_NEW_HTML, version=2
        )
        try:
            h = _h(seeded["u_a"])
            with TestClient(app) as client:
                # 差分 (v1 → v2): 実コンテンツ由来の行が現れる
                r = client.get(f"/outputs/{v2}/diff/{v1}", headers=h)
                assert r.status_code == 200
                d = r.json()["data"]
                assert (d["from_version"], d["to_version"]) == (1, 2)
                assert "-<p>旧仕様</p>" in d["diff"]
                assert "+<p>新仕様</p>" in d["diff"]
                assert d["added"] == 2 and d["removed"] == 1

                # 復元: v1 の内容が新版 v3 として積まれる (履歴は消えない)
                r2 = client.post(f"/outputs/{v1}/restore", headers=h)
                assert r2.status_code == 201
                restored = r2.json()["data"]
                assert restored["version"] == 3
                assert restored["meta"]["author"] == "restore"
                assert restored["meta"]["restored_from_version"] == 1
                # 復元版と v1 は同一内容 (identical) — 実体で確認
                r3 = client.get(f"/outputs/{restored['id']}/diff/{v1}", headers=h)
                assert r3.status_code == 200
                assert r3.json()["data"]["identical"] is True
                # 旧版はそのまま残る
                assert client.get(f"/outputs/{v1}", headers=h).status_code == 200

                # 最新版の復元は無意味な複製 — 409 で誠実に断る
                r4 = client.post(f"/outputs/{restored['id']}/restore", headers=h)
                assert r4.status_code == 409
                assert "最新版" in r4.json()["detail"]

                # 他 workspace ユーザーには不可視 (404)
                assert (
                    client.get(f"/outputs/{v2}/diff/{v1}", headers=_h(seeded["u_b"])).status_code
                    == 404
                )
                assert (
                    client.post(f"/outputs/{v1}/restore", headers=_h(seeded["u_b"])).status_code
                    == 404
                )
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.workflow_outputs "
                        "where project_id = cast(:p as uuid) and stage = 'requirements'"
                    ),
                    {"p": seeded["proj_a"]},
                )

    def test_diff_binary_and_no_html_409(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """バイナリ (filedb) はテキスト差分を偽装せず 409、本文なし復元も 409。"""
        test_engine = self._patch_service_factory(monkeypatch)
        proj = seeded["proj_a"]
        fn_meta = '{"file_name": "deck.pptx"}'
        b1 = _seed_output_version(
            sync_engine,
            project_id=proj,
            stage="design",
            html=None,
            path_override=f"filedb://{uuid.uuid4()}",
            version=1,
            meta=fn_meta,
        )
        b2 = _seed_output_version(
            sync_engine,
            project_id=proj,
            stage="design",
            html=None,
            path_override=f"filedb://{uuid.uuid4()}",
            version=2,
            meta=fn_meta,
        )
        try:
            h = _h(seeded["u_a"])
            with TestClient(app) as client:
                r = client.get(f"/outputs/{b2}/diff/{b1}", headers=h)
                assert r.status_code == 409
                # GAP-225: 「バイナリ」は利用者の言葉ではないので使わない
                assert "この形式のファイルは文章の差分を表示できません" in r.json()["detail"]
                # 本文を持たない成果物 (seeded の out_a) は復元対象なし → 409
                r2 = client.post(f"/outputs/{seeded['out_a']}/restore", headers=h)
                assert r2.status_code == 409
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.workflow_outputs "
                        "where project_id = cast(:p as uuid) and stage = 'design' "
                        "and id != cast(:keep as uuid)"
                    ),
                    {"p": seeded["proj_a"], "keep": seeded["out_a"]},
                )

    def test_anchors_work_for_mockdb_outputs(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """GAP-155 是正: mockdb 成果物の anchors が 503 にならず実 HTML から抽出される。"""
        test_engine = self._patch_service_factory(monkeypatch)
        v1 = _seed_output_version(
            sync_engine,
            project_id=seeded["proj_a"],
            stage="requirements",
            html='<html><body><h2 id="sec-1">1. 概要</h2><p>x</p></body></html>',
            version=1,
        )
        try:
            with TestClient(app) as client:
                r = client.get(f"/outputs/{v1}/anchors", headers=_h(seeded["u_a"]))
                assert r.status_code == 200
                assert r.json()["data"] == [{"element_id": "sec-1", "label": "1. 概要"}]
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.workflow_outputs "
                        "where project_id = cast(:p as uuid) and stage = 'requirements'"
                    ),
                    {"p": seeded["proj_a"]},
                )


# ── GAP-158: 出力デザインテンプレート (見た目の型・ワンダ生成・版連鎖) ────


def _patch_service_factory(monkeypatch: pytest.MonkeyPatch, engine: AsyncEngine) -> None:
    """service 経路 (mockdb / filedb / relay ジョブ) をテスト PG に向ける。

    本番は単一 event loop だが TestClient はブロック毎に新しい loop を作るため、
    NullPool のテスト engine に固定して loop 跨ぎの接続再利用を避ける。
    """
    from sqlalchemy.ext.asyncio import async_sessionmaker

    from src.services.mocks import artifacts as artifacts_svc
    from src.services.outputs import file_edit as file_edit_svc

    factory = async_sessionmaker(engine, class_=AsyncSession)
    monkeypatch.setattr(artifacts_svc, "service_session_factory", lambda: factory)
    monkeypatch.setattr(file_edit_svc, "job_session_factory", lambda: factory)


def _seed_design_template(
    sync_engine: sqlalchemy.Engine, *, ws: str, stage: str, html: str, version: int = 1
) -> str:
    """デザインテンプレ 1 版を直接 seed (mockdb 実体つき)。返り値 template id。"""
    with sync_engine.begin() as c:
        cid = c.execute(
            text("insert into public.mock_contents (html) values (:h) returning id"),
            {"h": html},
        ).scalar_one()
        return str(
            c.execute(
                text(
                    "insert into public.output_design_templates "
                    "(workspace_id, stage, version, html_storage_path, note) "
                    "values (cast(:w as uuid), cast(:s as workflow_stage_enum), :v, :p, 'seed') "
                    "returning id"
                ),
                {"w": ws, "s": stage, "v": version, "p": f"mockdb://{cid}"},
            ).scalar_one()
        )


@pytest.mark.integration
class TestGap158DesignTemplates:
    def test_wanda_creates_versions_preview_and_rls(
        self, app: FastAPI, seeded: dict[str, str], monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """作成 → 改訂 = 新版 → 版履歴 → 署名 URL プレビュー → RLS 越境 404。

        LLM は fake 経路 (ATELIER_ALLOW_FAKE_LLM=1) — relay/API 未設定の CI で
        「指示 → HTML 版が積まれる → プレビュー配信」の配線全体を検証する。
        """
        monkeypatch.setenv("ATELIER_ALLOW_FAKE_LLM", "1")
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        monkeypatch.delenv("ATELIER_LLM_PROVIDER", raising=False)
        ws, h = seeded["ws_a"], _h(seeded["u_a"])
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        try:
            with TestClient(app) as client:
                # v1 作成 (ワンダ生成 — fake) — note にサマリーが入る
                r = client.post(
                    f"/workspaces/{ws}/design-templates/estimate",
                    json={"instruction": "白基調でロゴ右上、明細は罫線ありの表"},
                    headers=h,
                )
                assert r.status_code == 201
                v1 = r.json()["data"]
                assert v1["version"] == 1 and v1["stage_label"] == "見積書"
                assert "デザインテンプレを更新" in v1["note"]
                # 改訂 = 新版 (履歴は消えない)
                r2 = client.post(
                    f"/workspaces/{ws}/design-templates/estimate",
                    json={"instruction": "ヘッダーを紺の帯に"},
                    headers=h,
                )
                assert r2.status_code == 201 and r2.json()["data"]["version"] == 2
                versions = client.get(
                    f"/workspaces/{ws}/design-templates/estimate/versions", headers=h
                ).json()["data"]
                assert [v["version"] for v in versions] == [2, 1]
                # 一覧は種類ごとの最新のみ
                listed = client.get(f"/workspaces/{ws}/design-templates", headers=h).json()["data"]
                assert [(t["stage"], t["version"]) for t in listed] == [("estimate", 2)]
                # 署名 URL → HTML 配信 (プレビュー iframe と同じ経路)。
                # v2 は改訂 (base_html あり) なので data-fake-revision を含む
                url = client.get(
                    f"/design-templates/{versions[0]['id']}/content-url", headers=h
                ).json()["data"]["url"]
                body = client.get(url)
                assert body.status_code == 200
                assert 'data-fake-template="1"' in body.text
                assert 'data-fake-revision="1"' in body.text
                # 未知の種類は 404 (偽の種類を作らない)
                assert (
                    client.post(
                        f"/workspaces/{ws}/design-templates/nonsense",
                        json={"instruction": "x"},
                        headers=h,
                    ).status_code
                    == 404
                )
                # RLS: 他人の workspace は不可視 (404)、テンプレ実体も 404
                hb = _h(seeded["u_b"])
                assert (
                    client.get(f"/workspaces/{ws}/design-templates", headers=hb).status_code == 404
                )
                assert (
                    client.get(
                        f"/design-templates/{versions[0]['id']}/content-url", headers=hb
                    ).status_code
                    == 404
                )
        finally:
            asyncio.run(test_engine.dispose())

    def test_llm_unavailable_503_no_fake_success(
        self, app: FastAPI, seeded: dict[str, str], monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """実行経路ゼロ (relay/サブスク/API 全て無し) は 503 — 偽テンプレを作らない。"""
        monkeypatch.delenv("ATELIER_ALLOW_FAKE_LLM", raising=False)
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        monkeypatch.delenv("ATELIER_LLM_PROVIDER", raising=False)
        with TestClient(app) as client:
            r = client.post(
                f"/workspaces/{seeded['ws_a']}/design-templates/estimate",
                json={"instruction": "紺基調で"},
                headers=_h(seeded["u_a"]),
            )
            assert r.status_code == 503
            versions = client.get(
                f"/workspaces/{seeded['ws_a']}/design-templates/estimate/versions",
                headers=_h(seeded["u_a"]),
            ).json()["data"]
            assert versions == []

    def test_design_injected_into_revise_system_prompt(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """スティーブ改訂の system prompt に該当種類のデザインテンプレが入る —
        「内容はスキル・指示、見た目はこのテンプレ」の分離が prompt で強制される。"""
        from typing import Any

        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        _seed_design_template(
            sync_engine,
            ws=seeded["ws_a"],
            stage="estimate",
            html="<html><body><h1>御見積書</h1><table class=meisai></table></body></html>",
        )
        oid = str(uuid.uuid4())
        with sync_engine.begin() as c:
            cid = c.execute(
                text("insert into public.mock_contents (html) values (:h) returning id"),
                {"h": "<html><body><p>見積本文</p></body></html>"},
            ).scalar_one()
            c.execute(
                text(
                    "insert into public.workflow_outputs "
                    "(id, project_id, stage, html_path, summary, version) "
                    "values (cast(:i as uuid), cast(:p as uuid), 'estimate', :path, '見積', 1)"
                ),
                {"i": oid, "p": seeded["proj_a"], "path": f"mockdb://{cid}"},
            )

        from src.services.outputs import revise as revise_svc

        captured: list[str] = []

        class _Capture:
            async def complete(self, **kwargs: Any) -> Any:
                captured.append(str(kwargs.get("system")))

                class _Res:
                    text = "<html><body><p>改訂済み</p></body></html>"

                return _Res()

        async def run() -> None:
            async with AsyncSession(test_engine) as session:
                created = await revise_svc.revise_output(
                    session,
                    actor_id=seeded["u_a"],
                    output_id=oid,
                    instruction="支払条件を強調して",
                    client=_Capture(),
                )
                assert created is not None
                await session.commit()

        try:
            asyncio.run(run())
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.workflow_outputs "
                        "where project_id = cast(:p as uuid) and stage = 'estimate'"
                    ),
                    {"p": seeded["proj_a"]},
                )
        assert len(captured) == 1
        assert "出力デザインテンプレート: 見積書（v1）" in captured[0]
        assert "内容 (テキスト・数値・項目) だけを依頼内容と" in captured[0]
        assert "デザインを勝手に変えないこと" in captured[0]
        assert "table class=meisai" in captured[0]

    def test_design_injected_into_chat_context_for_current_stage(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """チャット生成: 現在工程のデザインテンプレが system prompt に入る。

        GAP-159 で議事録は「デザインを持たない工程」になったため、現在工程を
        要件定義 (デザイン種類あり) に進めた状態で検証する。
        """
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        _seed_design_template(
            sync_engine,
            ws=seeded["ws_a"],
            stage="requirements",
            html="<html><body><h1>要件定義書</h1><ul class=agenda></ul></body></html>",
        )
        h = _h(seeded["u_a"])
        thread = str(uuid.uuid4())
        emp = str(uuid.uuid4())
        with sync_engine.begin() as c:
            # GAP-173: 運営シードが入った DB ではトリガが既に steve を作っている
            emp = ensure_ai_employee(
                c,
                workspace_id=seeded["ws_a"],
                name="steve",
                display_name="スティーブ",
                is_default=True,
                employee_id=emp,
            )
            c.execute(
                text(
                    "insert into public.chat_threads (id, project_id, ai_employee_id, title) "
                    "values (cast(:i as uuid), cast(:p as uuid), cast(:e as uuid), 'tmpl-t')"
                ),
                {"i": thread, "p": seeded["proj_a"], "e": emp},
            )
        try:
            with TestClient(app) as client:
                # フロー初期化 → 議事録を完了にして現在工程を要件定義へ進める
                client.get(f"/projects/{seeded['proj_a']}/flow", headers=h)
                with sync_engine.begin() as c:
                    c.execute(
                        text(
                            "update public.project_flow_stages set status = 'done' "
                            "where project_id = cast(:p as uuid) and stage_key = 'hearing'"
                        ),
                        {"p": seeded["proj_a"]},
                    )
                r = client.post(
                    f"/chat/threads/{thread}/context-preview",
                    headers=h,
                    json={"user_message": "要件をまとめて", "include_history": 5},
                )
                assert r.status_code == 200
                sys_p = r.json()["data"]["system_prompt"]
                assert "出力デザインテンプレート: 要件定義書（v1）" in sys_p
                assert "ul class=agenda" in sys_p
        finally:
            asyncio.run(test_engine.dispose())


# ── GAP-159: 運営既定デザイン (継承 / 上書き / 既定に戻す) ────


def _seed_platform_default(
    sync_engine: sqlalchemy.Engine, *, stage: str, html: str, version: int = 1
) -> str:
    with sync_engine.begin() as c:
        # 運営既定は全テナント共通のグローバル行 — 他テスト/実 e2e の残骸があると
        # 「継承しているのはこの既定」の検証が揺れるため、対象種類を掃除してから積む
        c.execute(
            text(
                "delete from public.output_design_templates "
                "where is_platform_default and stage = cast(:s as workflow_stage_enum)"
            ),
            {"s": stage},
        )
        cid = c.execute(
            text("insert into public.mock_contents (html) values (:h) returning id"),
            {"h": html},
        ).scalar_one()
        return str(
            c.execute(
                text(
                    "insert into public.output_design_templates "
                    "(workspace_id, stage, version, html_storage_path, note, "
                    " is_platform_default) "
                    "values (null, cast(:s as workflow_stage_enum), :v, :p, '運営既定', true) "
                    "returning id"
                ),
                {"s": stage, "v": version, "p": f"mockdb://{cid}"},
            ).scalar_one()
        )


@pytest.mark.integration
class TestGap159PlatformDefaults:
    def test_workspace_inherits_platform_default_until_it_makes_its_own(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """既定を継承 → WS が変更すると自前版が優先 → 既定に戻せる。"""
        monkeypatch.setenv("ATELIER_ALLOW_FAKE_LLM", "1")
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        monkeypatch.delenv("ATELIER_LLM_PROVIDER", raising=False)
        ws, h = seeded["ws_a"], _h(seeded["u_a"])
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        default_id = _seed_platform_default(
            sync_engine,
            stage="estimate",
            html="<html><body><h1 id='plat'>運営既定の見積書</h1></body></html>",
        )
        try:
            with TestClient(app) as client:
                # 1) 継承: WS は自前版ゼロでも既定が実効デザインとして返る
                listed = client.get(f"/workspaces/{ws}/design-templates", headers=h).json()["data"]
                est = [t for t in listed if t["stage"] == "estimate"]
                assert len(est) == 1
                assert est[0]["source"] == "platform" and est[0]["id"] == default_id
                assert est[0]["workspace_id"] is None
                # 自前の版履歴は空 (= 継承中)
                assert (
                    client.get(
                        f"/workspaces/{ws}/design-templates/estimate/versions", headers=h
                    ).json()["data"]
                    == []
                )
                # 既定 HTML はテナントからも閲覧できる (プレビュー用)
                url = client.get(f"/design-templates/{default_id}/content-url", headers=h).json()[
                    "data"
                ]["url"]
                assert "運営既定の見積書" in client.get(url).text

                # 2) 上書き: WS が変更すると自前版 v1 が実効になる (既定を土台に改訂)
                created = client.post(
                    f"/workspaces/{ws}/design-templates/estimate",
                    json={"instruction": "ロゴを右上に"},
                    headers=h,
                )
                assert created.status_code == 201
                assert created.json()["data"]["source"] == "workspace"
                listed2 = client.get(f"/workspaces/{ws}/design-templates", headers=h).json()["data"]
                est2 = next(t for t in listed2 if t["stage"] == "estimate")
                assert est2["source"] == "workspace" and est2["version"] == 1
                own_html_url = client.get(
                    f"/design-templates/{est2['id']}/content-url", headers=h
                ).json()["data"]["url"]
                own_html = client.get(own_html_url).text
                # 既定を土台に改訂したので fake の改訂マーカーが付く
                assert 'data-fake-revision="1"' in own_html

                # 3) 既定に戻す: 削除ではなく新版として積まれ、中身は既定と同一
                reset = client.post(
                    f"/workspaces/{ws}/design-templates/estimate/reset-to-default",
                    headers=h,
                )
                assert reset.status_code == 200
                back = reset.json()["data"]
                assert back["version"] == 2 and back["note"] == "運営の既定デザインに戻しました"
                back_url = client.get(
                    f"/design-templates/{back['id']}/content-url", headers=h
                ).json()["data"]["url"]
                assert "運営既定の見積書" in client.get(back_url).text
                # 履歴は消えない (v1 も残る)
                versions = client.get(
                    f"/workspaces/{ws}/design-templates/estimate/versions", headers=h
                ).json()["data"]
                assert [v["version"] for v in versions] == [2, 1]
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text("delete from public.output_design_templates where is_platform_default")
                )

    def test_platform_default_is_injected_when_workspace_has_none(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """WS 自前版が無くても、運営既定が AI の system prompt に注入される。"""
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        _seed_platform_default(
            sync_engine,
            stage="requirements",
            html="<html><body><h1>要件定義書</h1><ol class='req-list'></ol></body></html>",
        )
        h = _h(seeded["u_a"])
        thread, emp = str(uuid.uuid4()), str(uuid.uuid4())
        with sync_engine.begin() as c:
            # GAP-173: 運営シードが入った DB ではトリガが既に steve を作っている
            emp = ensure_ai_employee(
                c,
                workspace_id=seeded["ws_a"],
                name="steve",
                display_name="スティーブ",
                is_default=True,
                employee_id=emp,
            )
            c.execute(
                text(
                    "insert into public.chat_threads (id, project_id, ai_employee_id, title) "
                    "values (cast(:i as uuid), cast(:p as uuid), cast(:e as uuid), 'plat-t')"
                ),
                {"i": thread, "p": seeded["proj_a"], "e": emp},
            )
            # 現在工程を requirements にする (hearing を完了扱いにはせず直接 seed)
        try:
            with TestClient(app) as client:
                client.get(f"/projects/{seeded['proj_a']}/flow", headers=h)
                with sync_engine.begin() as c:
                    c.execute(
                        text(
                            "update public.project_flow_stages set status = 'done' "
                            "where project_id = cast(:p as uuid) and stage_key = 'hearing'"
                        ),
                        {"p": seeded["proj_a"]},
                    )
                r = client.post(
                    f"/chat/threads/{thread}/context-preview",
                    headers=h,
                    json={"user_message": "要件まとめて", "include_history": 5},
                )
                assert r.status_code == 200
                sys_p = r.json()["data"]["system_prompt"]
                assert "出力デザインテンプレート: 要件定義書（運営既定）" in sys_p
                assert "req-list" in sys_p
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text("delete from public.output_design_templates where is_platform_default")
                )

    def test_only_admin_can_write_platform_defaults(
        self, app: FastAPI, seeded: dict[str, str], monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """運営既定の作成は admin のみ (一般ユーザーは 403)。"""
        monkeypatch.setenv("ATELIER_ALLOW_FAKE_LLM", "1")
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        monkeypatch.delenv("ATELIER_LLM_PROVIDER", raising=False)
        with TestClient(app) as client:
            r = client.post(
                "/admin/design-templates/estimate",
                json={"instruction": "全社共通の見積デザイン"},
                headers=_h(seeded["u_a"]),
            )
            assert r.status_code == 403
            assert (
                client.get("/admin/design-templates", headers=_h(seeded["u_a"])).status_code == 403
            )

    def test_reference_files_are_injected_into_wanda_prompt(
        self, app: FastAPI, seeded: dict[str, str], monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """GAP-161: アップロードした参考資料の中身がテンプレ生成の prompt に入る。"""
        import io
        from typing import Any

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["区分", "金額"])
        ws.append(["月額保守", 120000])
        buf = io.BytesIO()
        wb.save(buf)
        payload = buf.getvalue()

        from src import storage_signing

        monkeypatch.setenv("ATELIER_SUPABASE_ADMIN_API_URL", "https://stor.invalid")
        monkeypatch.setenv("ATELIER_SUPABASE_SERVICE_ROLE_KEY", "svc-key")
        monkeypatch.setenv("ATELIER_ALLOW_FAKE_LLM", "1")
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        monkeypatch.delenv("ATELIER_LLM_PROVIDER", raising=False)

        class _Res:
            def __init__(self) -> None:
                self.status_code = 200
                self.content = payload

            def json(self) -> dict[str, Any]:
                return {"signedURL": "/object/sign/ref/x?token=d"}

        class _Client:
            def __init__(self, *_a: Any, **_k: Any) -> None: ...

            async def __aenter__(self) -> _Client:
                return self

            async def __aexit__(self, *_a: Any) -> bool:
                return False

            async def post(self, _url: str, **_k: Any) -> _Res:
                return _Res()

            async def get(self, _url: str, **_k: Any) -> _Res:
                return _Res()

        monkeypatch.setattr(storage_signing.httpx, "AsyncClient", _Client)
        import httpx as _httpx

        monkeypatch.setattr(_httpx, "AsyncClient", _Client)

        captured: list[str] = []
        from src.services.chat_sse import llm_chain

        async def _fake_complete(
            *, system_prompt: str, user_text: str, actor_id: str, **_k: Any
        ) -> tuple[str, str]:
            del user_text, actor_id
            captured.append(system_prompt)
            return "<!doctype html><html><body>ok</body></html>\n---SUMMARY---\nテスト", "fake"

        monkeypatch.setattr(llm_chain, "llm_complete", _fake_complete)

        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        try:
            with TestClient(app) as client:
                r = client.post(
                    f"/workspaces/{seeded['ws_a']}/design-templates/invoice",
                    headers=_h(seeded["u_a"]),
                    json={
                        "instruction": "この請求書の様式に寄せて",
                        "reference_files": [
                            {
                                "storage_path": "reference-uploads/u/請求例.xlsx",
                                "file_name": "請求例.xlsx",
                                "mime_type": (
                                    "application/vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet"
                                ),
                            }
                        ],
                    },
                )
                assert r.status_code == 201, r.text
        finally:
            asyncio.run(test_engine.dispose())
        assert len(captured) == 1
        assert "# 参考資料" in captured[0]
        assert "請求例.xlsx" in captured[0]
        assert "月額保守 | 120000" in captured[0]


# ── GAP-162: クライアント共有リンク + 書き出し (HTML / Excel) ────


@pytest.mark.integration
class TestGap162ShareAndExport:
    def test_gap324_325_export_is_audited_and_viewer_cannot_ai_edit(
        self, app: FastAPI, seeded: dict[str, str], sync_engine: sqlalchemy.Engine
    ) -> None:
        """GAP-324: 書き出しは監査に残る (output.downloaded)。
        GAP-325: 「AI に修正を依頼」も権限判定が Bridge 判定より先 (viewer は 403)。"""
        oid = str(uuid.uuid4())
        with sync_engine.begin() as c:
            cid = c.execute(
                text("insert into public.mock_contents (html) values (:h) returning id"),
                {"h": "<html><body><table><tr><td>A</td></tr></table></body></html>"},
            ).scalar_one()
            c.execute(
                text(
                    "insert into public.workflow_outputs "
                    "(id, project_id, stage, html_path, summary, version) "
                    "values (cast(:i as uuid), cast(:p as uuid), 'estimate', :path, '御見積書', 1)"
                ),
                {"i": oid, "p": seeded["proj_a"], "path": f"mockdb://{cid}"},
            )
            c.execute(
                text(
                    "insert into public.workspace_memberships (workspace_id, user_id, role) "
                    "values (cast(:w as uuid), cast(:u as uuid), 'viewer') on conflict do nothing"
                ),
                {"w": seeded["ws_a"], "u": seeded["u_b"]},
            )
        try:
            with TestClient(app) as client:
                r = client.get(f"/outputs/{oid}/export?format=html", headers=_h(seeded["u_a"]))
                assert r.status_code == 200, r.text
                # GAP-325: 閲覧者の「AI に修正を依頼」は 503 (未接続) ではなく 403
                denied = client.post(
                    f"/outputs/{oid}/ai-file-edit",
                    json={"instruction": "直して"},
                    headers=_h(seeded["u_b"]),
                )
                assert denied.status_code == 403, denied.text
                assert "権限" in denied.json()["detail"]
            with sync_engine.begin() as c:
                rows = c.execute(
                    text(
                        "select after from public.audit_logs "
                        "where action = 'output.downloaded' and target_id = :i"
                    ),
                    {"i": oid},
                ).all()
                assert len(rows) == 1, rows
                after = (
                    rows[0].after if isinstance(rows[0].after, dict) else json.loads(rows[0].after)
                )
                assert after["format"] == "html"
        finally:
            with sync_engine.begin() as c:
                c.execute(text("delete from public.audit_logs where target_id = :i"), {"i": oid})
                c.execute(
                    text("delete from public.workflow_outputs where id = cast(:i as uuid)"),
                    {"i": oid},
                )
                c.execute(text("delete from public.mock_contents where id = :c"), {"c": cid})
                c.execute(
                    text(
                        "delete from public.workspace_memberships where workspace_id = cast(:w as uuid) "
                        "and user_id = cast(:u as uuid)"
                    ),
                    {"w": seeded["ws_a"], "u": seeded["u_b"]},
                )

    def test_share_link_roundtrip_and_revoke(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """発行 → 認証なしで閲覧できる → 失効すると 410。"""
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = str(uuid.uuid4())
        with sync_engine.begin() as c:
            cid = c.execute(
                text("insert into public.mock_contents (html) values (:h) returning id"),
                {
                    "h": "<html><body><h1>御見積書</h1>"
                    "<table><tr><th>項目</th><th>金額</th></tr>"
                    "<tr><td>設計</td><td>400000</td></tr></table></body></html>"
                },
            ).scalar_one()
            c.execute(
                text(
                    "insert into public.workflow_outputs "
                    "(id, project_id, stage, html_path, summary, version) "
                    "values (cast(:i as uuid), cast(:p as uuid), 'estimate', :path, '御見積書', 1)"
                ),
                {"i": oid, "p": seeded["proj_a"], "path": f"mockdb://{cid}"},
            )
        h = _h(seeded["u_a"])
        try:
            with TestClient(app) as client:
                # GAP-307: 閲覧者 (viewer) は共有リンクを発行できない (以前は 201 で発行できた)
                with sync_engine.begin() as c:
                    c.execute(
                        text(
                            "insert into public.workspace_memberships (workspace_id, user_id, role) "
                            "values (cast(:w as uuid), cast(:u as uuid), 'viewer') on conflict do nothing"
                        ),
                        {"w": seeded["ws_a"], "u": seeded["u_b"]},
                    )
                denied = client.post(
                    f"/outputs/{oid}/share-links",
                    headers=_h(seeded["u_b"]),
                    json={"label": "viewer が発行", "expires_days": 7},
                )
                assert denied.status_code == 403, denied.text
                # GAP-311 (通し J46-18 再測): revise / restore も権限判定が先 (503 / 409 より前に 403)
                assert (
                    client.post(
                        f"/outputs/{oid}/revise",
                        headers=_h(seeded["u_b"]),
                        json={"instruction": "viewer が改訂"},
                    ).status_code
                    == 403
                )
                assert (
                    client.post(f"/outputs/{oid}/restore", headers=_h(seeded["u_b"])).status_code
                    == 403
                )
                with sync_engine.begin() as c:
                    c.execute(
                        text(
                            "delete from public.workspace_memberships where workspace_id = cast(:w as uuid) "
                            "and user_id = cast(:u as uuid)"
                        ),
                        {"w": seeded["ws_a"], "u": seeded["u_b"]},
                    )
                created = client.post(
                    f"/outputs/{oid}/share-links",
                    headers=h,
                    json={"label": "田中様へ", "expires_days": 7},
                )
                assert created.status_code == 201, created.text
                link = created.json()["data"]
                assert link["label"] == "田中様へ"
                url = link["share_url"]
                assert "/share/" in url
                token = url.rsplit("/share/", 1)[1]

                # 認証なしで閲覧できる (クライアントに渡すため)
                viewed = client.get(f"/share/{token}")
                assert viewed.status_code == 200
                assert "御見積書" in viewed.text
                assert "PDF で保存 / 印刷" in viewed.text  # PDF 化の導線

                # 一覧では URL 自体は返らない (ハッシュしか保存していない)
                listed = client.get(f"/outputs/{oid}/share-links", headers=h).json()["data"]
                assert len(listed) == 1 and listed[0]["share_url"] is None
                assert listed[0]["view_count"] >= 1

                # 失効 → 以後は 410
                rev = client.post(f"/share-links/{link['id']}/revoke", headers=h)
                assert rev.status_code == 200
                assert client.get(f"/share/{token}").status_code == 410
                # 二重失効は 404
                assert (
                    client.post(f"/share-links/{link['id']}/revoke", headers=h).status_code == 404
                )
                # 存在しないトークンは 404
                assert client.get("/share/nonexistent-token").status_code == 404

                # GAP-319 (通し R3 所見 / G-14): **案件を削除したら共有リンクも辿れない**。
                # 以前は workflow_outputs.deleted_at しか見ておらず、案件・WS を消しても
                # 外部の公開 URL が 200 で中身を返し続けていた。
                link2 = client.post(
                    f"/outputs/{oid}/share-links",
                    headers=h,
                    json={"label": "案件削除の検証", "expires_days": 7},
                ).json()["data"]
                token2 = str(link2["share_url"]).rsplit("/share/", 1)[1]
                assert client.get(f"/share/{token2}").status_code == 200
                assert client.delete(f"/projects/{seeded['proj_a']}", headers=h).status_code == 204
                assert client.get(f"/share/{token2}").status_code == 404
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text("delete from public.workflow_outputs where id = cast(:i as uuid)"),
                    {"i": oid},
                )

    def test_export_html_and_xlsx_and_honest_refusal(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """HTML はそのまま、表つきは Excel、表が無ければ正直に 409。"""
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        with_table, no_table = str(uuid.uuid4()), str(uuid.uuid4())
        with sync_engine.begin() as c:
            c1 = c.execute(
                text("insert into public.mock_contents (html) values (:h) returning id"),
                {
                    "h": "<html><body><table><tr><th>項目</th><th>金額</th></tr>"
                    "<tr><td>実装</td><td>1200000</td></tr></table></body></html>"
                },
            ).scalar_one()
            c2 = c.execute(
                text("insert into public.mock_contents (html) values (:h) returning id"),
                {"h": "<html><body><p>表のない提案書です</p></body></html>"},
            ).scalar_one()
            for oid, cid, stage, summary in (
                (with_table, c1, "estimate", "見積書"),
                (no_table, c2, "proposal", "提案書"),
            ):
                c.execute(
                    text(
                        "insert into public.workflow_outputs "
                        "(id, project_id, stage, html_path, summary, version) values "
                        "(cast(:i as uuid), cast(:p as uuid), cast(:s as workflow_stage_enum), "
                        " :path, :sm, 1)"
                    ),
                    {
                        "i": oid,
                        "p": seeded["proj_a"],
                        "s": stage,
                        "path": f"mockdb://{cid}",
                        "sm": summary,
                    },
                )
        h = _h(seeded["u_a"])
        try:
            with TestClient(app) as client:
                r_html = client.get(f"/outputs/{with_table}/export?format=html", headers=h)
                assert r_html.status_code == 200
                assert "attachment" in r_html.headers["content-disposition"]
                assert "1200000" in r_html.text

                r_xlsx = client.get(f"/outputs/{with_table}/export?format=xlsx", headers=h)
                assert r_xlsx.status_code == 200
                assert r_xlsx.headers["content-type"].startswith(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                # 実 xlsx として開けて、中身が入っていること
                import io

                from openpyxl import load_workbook

                wb = load_workbook(io.BytesIO(r_xlsx.content))
                rows = [list(r) for r in wb.worksheets[0].iter_rows(values_only=True)]
                assert list(rows[0]) == ["項目", "金額"]
                assert list(rows[1]) == ["実装", "1200000"]

                # 表が無い成果物は「できない」と正直に断る (空ファイルを出さない)
                bad = client.get(f"/outputs/{no_table}/export?format=xlsx", headers=h)
                assert bad.status_code == 409
                assert "表が無い" in bad.json()["detail"]
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.workflow_outputs where id in (cast(:a as uuid), cast(:b as uuid))"
                    ),
                    {"a": with_table, "b": no_table},
                )


# ── GAP-163: Excel / CSV の表表示と編集 ────────────────────────────


@pytest.mark.integration
class TestGap163SheetViewAndEdit:
    def _seed_xlsx_output(self, sync_engine: sqlalchemy.Engine, project_id: str) -> str:
        import io

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "明細"
        ws.append(["項目", "金額"])
        ws.append(["設計", "300000"])
        buf = io.BytesIO()
        wb.save(buf)
        oid = str(uuid.uuid4())
        with sync_engine.begin() as c:
            fid = c.execute(
                text(
                    "insert into public.artifact_files (data, mime, file_name, byte_size) "
                    "values (:d, :m, :n, :s) returning id"
                ),
                {
                    "d": buf.getvalue(),
                    "m": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "n": "見積明細.xlsx",
                    "s": len(buf.getvalue()),
                },
            ).scalar_one()
            c.execute(
                text(
                    "insert into public.workflow_outputs "
                    "(id, project_id, stage, html_path, summary, version) values "
                    "(cast(:i as uuid), cast(:p as uuid), 'estimate', :path, '見積明細', 1)"
                ),
                {"i": oid, "p": project_id, "path": f"filedb://{fid}"},
            )
        return oid

    def test_excel_is_shown_as_table_and_edit_creates_new_version(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = self._seed_xlsx_output(sync_engine, seeded["proj_a"])
        h = _h(seeded["u_a"])
        try:
            with TestClient(app) as client:
                got = client.get(f"/outputs/{oid}/sheet", headers=h)
                assert got.status_code == 200, got.text
                data = got.json()["data"]
                assert data["editable"] is True
                assert data["file_name"] == "見積明細.xlsx"
                assert data["sheets"][0]["name"] == "明細"
                assert data["sheets"][0]["rows"][0] == ["項目", "金額"]
                assert data["sheets"][0]["rows"][1] == ["設計", "300000"]
                # 保持しないもの (数式・書式) を正直に伝えている
                assert "数式・書式" in data["note"]

                # セルを編集して保存 → 新バージョン (元の版は残る)
                saved = client.post(
                    f"/outputs/{oid}/sheet",
                    headers=h,
                    json={
                        "sheets": [
                            {
                                "name": "明細",
                                "rows": [["項目", "金額"], ["設計", "450000"], ["実装", "800000"]],
                            }
                        ]
                    },
                )
                assert saved.status_code == 201, saved.text
                new_id = saved.json()["data"]["id"]
                assert saved.json()["data"]["version"] == 2

                # 新版を読み直すと編集が反映されている
                after = client.get(f"/outputs/{new_id}/sheet", headers=h).json()["data"]
                assert after["sheets"][0]["rows"][1] == ["設計", "450000"]
                assert after["sheets"][0]["rows"][2] == ["実装", "800000"]
                # 元の版は変わらない (履歴不滅)
                before = client.get(f"/outputs/{oid}/sheet", headers=h).json()["data"]
                assert before["sheets"][0]["rows"][1] == ["設計", "300000"]
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.workflow_outputs where project_id = cast(:p as uuid) "
                        "and summary = '見積明細'"
                    ),
                    {"p": seeded["proj_a"]},
                )

    def test_gap254_stale_base_version_is_409_not_silent_overwrite(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """GAP-254 (本番実走 SG01-217): 2 つのタブで同じ表を編集し両方保存 → 両方 201 で、
        相手の編集を含まない古い内容が最新版になっていた。基底版 (GET /sheet の version) を
        添えて保存し、最新と違えば 409 で止める。"""
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = self._seed_xlsx_output(sync_engine, seeded["proj_a"])
        h = _h(seeded["u_a"])
        rows_a = [["項目", "金額"], ["設計", "111"]]
        rows_b = [["項目", "金額"], ["設計", "222"]]
        try:
            with TestClient(app) as client:
                got = client.get(f"/outputs/{oid}/sheet", headers=h).json()["data"]
                assert got["version"] == 1  # 編集の基底版を画面に渡す
                # タブ 1: v1 を基底に保存 → v2
                r1 = client.post(
                    f"/outputs/{oid}/sheet",
                    headers=h,
                    json={"sheets": [{"name": "明細", "rows": rows_a}], "base_version": 1},
                )
                assert r1.status_code == 201, r1.text
                assert r1.json()["data"]["version"] == 2
                # タブ 2: まだ v1 を基底にしたまま保存 → 409 (黙って v3 にしない)
                r2 = client.post(
                    f"/outputs/{oid}/sheet",
                    headers=h,
                    json={"sheets": [{"name": "明細", "rows": rows_b}], "base_version": 1},
                )
                assert r2.status_code == 409, r2.text
                assert "先に新しい版を保存" in r2.json()["detail"], r2.text
                # 開き直して最新 (v2) を基底にすれば保存できる
                r3 = client.post(
                    f"/outputs/{oid}/sheet",
                    headers=h,
                    json={"sheets": [{"name": "明細", "rows": rows_b}], "base_version": 2},
                )
                assert r3.status_code == 201, r3.text
                assert r3.json()["data"]["version"] == 3
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.workflow_outputs where project_id = cast(:p as uuid) "
                        "and summary = '見積明細'"
                    ),
                    {"p": seeded["proj_a"]},
                )

    def test_gap265_new_version_notifies_invited_clients(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """GAP-265 (通し J21-05): 成果物に新版が積まれたら、招待済み (未失効) のクライアントへ
        更新メールを送り、監査ログ client_notified_of_update を残す。失効した招待には送らない。"""
        monkeypatch.delenv("ATELIER_EMAIL_API_KEY", raising=False)  # dry-run
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = self._seed_xlsx_output(sync_engine, seeded["proj_a"])
        h = _h(seeded["u_a"])
        try:
            with TestClient(app) as client:
                inv = client.post(
                    "/client-invitations",
                    json={"project_id": seeded["proj_a"], "email": "notify@ext.example"},
                    headers=h,
                )
                assert inv.status_code == 201, inv.text
                revoked = client.post(
                    "/client-invitations",
                    json={"project_id": seeded["proj_a"], "email": "revoked@ext.example"},
                    headers=h,
                ).json()["data"]["id"]
                assert (
                    client.post(f"/client-invitations/{revoked}/revoke", headers=h).status_code
                    == 200
                )
                saved = client.post(
                    f"/outputs/{oid}/sheet",
                    headers=h,
                    json={"sheets": [{"name": "明細", "rows": [["項目", "金額"], ["設計", "1"]]}]},
                )
                assert saved.status_code == 201, saved.text
                new_id = saved.json()["data"]["id"]
            with sync_engine.connect() as c:
                rows = c.execute(
                    text(
                        "select after from public.audit_logs where action='output.client_notified_of_update' "
                        "and target_id = cast(:t as uuid)"
                    ),
                    {"t": new_id},
                ).all()
            assert len(rows) == 1, rows  # 失効した招待には送らない
            after = rows[0][0] if isinstance(rows[0][0], dict) else json.loads(rows[0][0])
            assert after["version"] == 2
            assert after["dry_run"] is True
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.workflow_outputs where project_id = cast(:p as uuid) "
                        "and summary = '見積明細'"
                    ),
                    {"p": seeded["proj_a"]},
                )

    def test_pdf_is_viewable_but_edit_is_honestly_refused(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """PDF は表示できるが編集はできない — できないことを正直に言う。"""
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = str(uuid.uuid4())
        with sync_engine.begin() as c:
            fid = c.execute(
                text(
                    "insert into public.artifact_files (data, mime, file_name, byte_size) "
                    "values (:d, :m, :n, :s) returning id"
                ),
                {"d": b"%PDF-1.4 fake", "m": "application/pdf", "n": "契約書.pdf", "s": 13},
            ).scalar_one()
            c.execute(
                text(
                    "insert into public.workflow_outputs "
                    "(id, project_id, stage, html_path, summary, version) values "
                    "(cast(:i as uuid), cast(:p as uuid), 'contract', :path, '契約書', 1)"
                ),
                {"i": oid, "p": seeded["proj_a"], "path": f"filedb://{fid}"},
            )
        h = _h(seeded["u_a"])
        try:
            with TestClient(app) as client:
                r = client.get(f"/outputs/{oid}/sheet", headers=h)
                assert r.status_code == 409
                detail = r.json()["detail"]
                # GAP-225: PDF だけの案内は他の「表として扱えない」形式と混ぜない
                # (混ぜると「どう直すか」が消える)。code を分けて表から引く。
                assert "PDF はこの画面で表示できます" in detail
                assert "直接の編集はできません" in detail
                # 表示自体は inline 配信で可能 (content-url → content)
                url = client.get(f"/outputs/{oid}/content-url", headers=h).json()["data"]["url"]
                served = client.get(url)
                assert served.status_code == 200
                assert served.headers["content-type"].startswith("application/pdf")
                assert served.headers["content-disposition"].startswith("inline")
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text("delete from public.workflow_outputs where id = cast(:i as uuid)"),
                    {"i": oid},
                )


# ── GAP-166: ファイル成果物を本人の Claude Code に直してもらう ────────


@pytest.mark.integration
class TestGap166AiFileEdit:
    def _seed_file_output(
        self, sync_engine: sqlalchemy.Engine, project_id: str, *, name: str, mime: str
    ) -> str:
        oid = str(uuid.uuid4())
        with sync_engine.begin() as c:
            fid = c.execute(
                text(
                    "insert into public.artifact_files (data, mime, file_name, byte_size) "
                    "values (:d, :m, :n, :s) returning id"
                ),
                {"d": b"binary-content", "m": mime, "n": name, "s": 14},
            ).scalar_one()
            c.execute(
                text(
                    "insert into public.workflow_outputs "
                    "(id, project_id, stage, html_path, summary, version) values "
                    "(cast(:i as uuid), cast(:p as uuid), 'estimate', :path, :sm, 1)"
                ),
                {"i": oid, "p": project_id, "path": f"filedb://{fid}", "sm": name},
            )
        return oid

    def test_bridge_offline_is_refused_without_queueing(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Bridge がオフラインなら 503 で断る (黙ってキューに積んで待たせない)。"""
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = self._seed_file_output(
            sync_engine, seeded["proj_a"], name="見積.xlsx", mime="application/vnd.ms-excel"
        )
        emp, thread = str(uuid.uuid4()), str(uuid.uuid4())
        with sync_engine.begin() as c:
            # GAP-173: 運営シードが入った DB ではトリガが既に steve を作っている
            emp = ensure_ai_employee(
                c,
                workspace_id=seeded["ws_a"],
                name="steve",
                display_name="スティーブ",
                is_default=True,
                employee_id=emp,
            )
            c.execute(
                text(
                    "insert into public.chat_threads (id, project_id, ai_employee_id, title) "
                    "values (cast(:i as uuid), cast(:p as uuid), cast(:e as uuid), 'main')"
                ),
                {"i": thread, "p": seeded["proj_a"], "e": emp},
            )
            c.execute(text("delete from public.bridge_workers"))
        try:
            with TestClient(app) as client:
                r = client.post(
                    f"/outputs/{oid}/ai-file-edit",
                    headers=_h(seeded["u_a"]),
                    json={"instruction": "単価を 10% 上げて"},
                )
            assert r.status_code == 503
            assert "Bridge" in r.json()["detail"]
            # ジョブは積まれていない (この利用者の分で数える)
            with sync_engine.begin() as c:
                n = c.execute(
                    text(
                        "select count(*) from public.chat_relay_jobs "
                        "where requested_by = cast(:u as uuid)"
                    ),
                    {"u": seeded["u_a"]},
                ).scalar_one()
            assert n == 0
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text("delete from public.workflow_outputs where id = cast(:i as uuid)"),
                    {"i": oid},
                )

    def test_pdf_edit_is_queued_for_the_users_own_claude_code(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """PDF もジョブとして本人の PC へ渡る (サーバーが諦めるのではなく Claude Code が直す)。"""
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = self._seed_file_output(
            sync_engine, seeded["proj_a"], name="契約書.pdf", mime="application/pdf"
        )
        emp, thread, worker = str(uuid.uuid4()), str(uuid.uuid4()), str(uuid.uuid4())
        with sync_engine.begin() as c:
            # GAP-173: 運営シードが入った DB ではトリガが既に steve を作っている
            emp = ensure_ai_employee(
                c,
                workspace_id=seeded["ws_a"],
                name="steve",
                display_name="スティーブ",
                is_default=True,
                employee_id=emp,
            )
            c.execute(
                text(
                    "insert into public.chat_threads (id, project_id, ai_employee_id, title) "
                    "values (cast(:i as uuid), cast(:p as uuid), cast(:e as uuid), 'main')"
                ),
                {"i": thread, "p": seeded["proj_a"], "e": emp},
            )
            # Bridge をオンライン扱いにする
            c.execute(
                text(
                    "insert into public.bridge_workers (id, host_label, version, last_seen_at) "
                    "values (:i, 'mac', '1.0.0', now())"
                ),
                {"i": worker},
            )
        try:
            with TestClient(app) as client:
                r = client.post(
                    f"/outputs/{oid}/ai-file-edit",
                    headers=_h(seeded["u_a"]),
                    json={"instruction": "第 3 条の支払期日を月末締め翌月末に直して"},
                )
                assert r.status_code == 202, r.text
                job_id = r.json()["data"]["job_id"]
            with sync_engine.begin() as c:
                job = c.execute(
                    text(
                        "select prompt, system_prompt, tools_mode, requested_by, status "
                        "from public.chat_relay_jobs where id = cast(:i as uuid)"
                    ),
                    {"i": job_id},
                ).one()
            # 本人の PC で・ファイル名を指定して・同名保存を指示している
            assert "契約書.pdf" in job.prompt
            assert "第 3 条の支払期日" in job.prompt
            assert "同じファイル名で上書き保存" in job.prompt
            assert "形式は変えない" in job.system_prompt
            assert job.tools_mode == "auto"
            assert str(job.requested_by) == seeded["u_a"]
            assert job.status == "queued"
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text(
                        "delete from public.chat_relay_jobs where requested_by = cast(:u as uuid)"
                    ),
                    {"u": seeded["u_a"]},
                )
                c.execute(text("delete from public.bridge_workers where id = :i"), {"i": worker})
                c.execute(
                    text("delete from public.workflow_outputs where id = cast(:i as uuid)"),
                    {"i": oid},
                )

    def test_file_is_delivered_to_the_working_folder(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """作業場 seed に対象ファイルの実体 (base64) が含まれる = Claude Code が開ける。"""
        from src.services.chat_relay import get_job_workspace_seed

        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = self._seed_file_output(
            sync_engine,
            seeded["proj_a"],
            name="見積明細.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        emp, thread, job = str(uuid.uuid4()), str(uuid.uuid4()), str(uuid.uuid4())
        with sync_engine.begin() as c:
            # GAP-173: 運営シードが入った DB ではトリガが既に steve を作っている
            emp = ensure_ai_employee(
                c,
                workspace_id=seeded["ws_a"],
                name="steve",
                display_name="スティーブ",
                is_default=True,
                employee_id=emp,
            )
            c.execute(
                text(
                    "insert into public.chat_threads (id, project_id, ai_employee_id, title) "
                    "values (cast(:i as uuid), cast(:p as uuid), cast(:e as uuid), 'main')"
                ),
                {"i": thread, "p": seeded["proj_a"], "e": emp},
            )
            c.execute(
                text(
                    "insert into public.chat_relay_jobs "
                    "(id, thread_id, requested_by, status, system_prompt, prompt, tools_mode) "
                    "values (cast(:i as uuid), cast(:t as uuid), cast(:u as uuid), 'running', "
                    "        's', 'p', 'auto')"
                ),
                {"i": job, "t": thread, "u": seeded["u_a"]},
            )

        async def run() -> list[dict[str, str]]:
            engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
            try:
                async with AsyncSession(engine) as session:
                    return await get_job_workspace_seed(session, job_id=job)
            finally:
                await engine.dispose()

        try:
            seed = asyncio.run(run())
            names = [f["file_name"] for f in seed]
            assert "見積明細.xlsx" in names
            target = next(f for f in seed if f["file_name"] == "見積明細.xlsx")
            # HTML ではなく実体 (base64) で配られる
            assert target.get("content_b64")
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text("delete from public.chat_relay_jobs where id = cast(:i as uuid)"),
                    {"i": job},
                )
                c.execute(
                    text("delete from public.workflow_outputs where id = cast(:i as uuid)"),
                    {"i": oid},
                )


@pytest.mark.integration
class TestGap171SubscriptionRouting:
    """GAP-171: スティーブ (成果物の改訂・修正提案) も本人の Claude サブスクで動く。

    以前は `ANTHROPIC_API_KEY` (= 運営の従量課金) を直接叩いており、
    「全ユーザーが自分の PC・自分のサブスクで実行する」という確定アーキテクチャと
    食い違っていた。ここでは **運営のキーを一切設定しない**状態で:
      - relay 指定 + Bridge オフライン → 503 (bridge_offline) で正直に断り、版は積まない
      - 経路がまったく無い → 503 (unconfigured)
    を実測する。偽の改訂を運営費用で作らないことの担保。
    """

    def _seed_output(self, sync_engine: sqlalchemy.Engine, project_id: str) -> str:
        oid = str(uuid.uuid4())
        with sync_engine.begin() as c:
            cid = c.execute(
                text("insert into public.mock_contents (html) values (:h) returning id"),
                {"h": "<html><body><h1>GAP-171</h1></body></html>"},
            ).scalar_one()
            c.execute(
                text(
                    "insert into public.workflow_outputs "
                    "(id, project_id, stage, html_path, summary, version) values "
                    "(cast(:i as uuid), cast(:p as uuid), 'requirements', "
                    ":path, 'GAP-171 検証', 1)"
                ),
                {"i": oid, "p": project_id, "path": f"mockdb://{cid}"},
            )
        return oid

    def _versions(self, sync_engine: sqlalchemy.Engine, project_id: str) -> int:
        with sync_engine.begin() as c:
            return int(
                c.execute(
                    text(
                        "select count(*) from public.workflow_outputs "
                        "where project_id = cast(:p as uuid) and deleted_at is null"
                    ),
                    {"p": project_id},
                ).scalar_one()
            )

    def test_revise_uses_bridge_and_refuses_honestly_when_offline(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """relay 指定 + Bridge オフライン → 503。運営キーへ黙って落ちない。"""
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        monkeypatch.delenv("ATELIER_ALLOW_FAKE_LLM", raising=False)
        monkeypatch.setenv("ATELIER_LLM_PROVIDER", "relay")
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = self._seed_output(sync_engine, seeded["proj_a"])
        before = self._versions(sync_engine, seeded["proj_a"])
        with sync_engine.begin() as c:
            c.execute(text("delete from public.bridge_workers"))
        try:
            with TestClient(app) as client:
                r = client.post(
                    f"/outputs/{oid}/revise",
                    headers=_h(seeded["u_a"]),
                    json={"instruction": "2 章を詳しく"},
                )
            assert r.status_code == 503, r.text
            assert "Bridge" in r.json()["detail"]
            # 偽の新版を積まない
            assert self._versions(sync_engine, seeded["proj_a"]) == before
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text("delete from public.workflow_outputs where id = cast(:i as uuid)"),
                    {"i": oid},
                )

    def test_revise_without_any_route_is_503_not_owner_billed(
        self,
        app: FastAPI,
        seeded: dict[str, str],
        sync_engine: sqlalchemy.Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """経路ゼロ (運営キーも無い) → 503。運営費用で勝手に生成しない。"""
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        monkeypatch.delenv("ATELIER_ALLOW_FAKE_LLM", raising=False)
        monkeypatch.delenv("ATELIER_LLM_PROVIDER", raising=False)
        test_engine = create_async_engine(PG_ASYNC, poolclass=NullPool)
        _patch_service_factory(monkeypatch, test_engine)
        oid = self._seed_output(sync_engine, seeded["proj_a"])
        before = self._versions(sync_engine, seeded["proj_a"])
        try:
            with TestClient(app) as client:
                r = client.post(
                    f"/outputs/{oid}/revise",
                    headers=_h(seeded["u_a"]),
                    json={"instruction": "2 章を詳しく"},
                )
            assert r.status_code == 503, r.text
            assert self._versions(sync_engine, seeded["proj_a"]) == before
        finally:
            asyncio.run(test_engine.dispose())
            with sync_engine.begin() as c:
                c.execute(
                    text("delete from public.workflow_outputs where id = cast(:i as uuid)"),
                    {"i": oid},
                )
