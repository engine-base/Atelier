#!/usr/bin/env python3
"""通しフロー Excel を「読める形」で生成する (GAP-214)。

なぜ作り直したか
----------------
初版は 82 行のフラットな 1 シートで、**フローがフローとして見えなかった**。
「この行はどの流れの何番目か」「この流れは何件中何件終わったか」が読み取れず、
経営者から「見せ方が悪い・わかりづらい」と指摘された。

Excel は記録の器であると同時に **人が読む資料**でもある。読めない資料は、
結局また「手元で別のものを作る」ことになり、正本が二重化する（鉄則3 の敵）。

構成
----
  フロー一覧   全フロー (43 本) を 1 行ずつ。進捗と状態が一目で分かる（**最初に開くシート**）
  第0群〜第6群 群ごとのタブ。フロー名の見出し行 → その配下にステップ
  Plan         機械可読（journey_workbook.py の update/status がここを読む）
  Roles        ロールと依存
  Summary      全体の集計と DONE? 判定

使い方
------
    python3 render_xlsx.py --plan plan.json --out ../../../../.qa/e2e-journey/journey-<日付>.xlsx
    # 実行結果 (status) を既存 Excel から引き継いで再生成する場合:
    python3 render_xlsx.py --plan plan.json --out <同じパス> --carry-status
"""

from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 見た目: 淡い色。印刷しても潰れない濃さにする。
C_HEAD = PatternFill("solid", fgColor="2F4858")  # 見出し行
C_FLOW = PatternFill("solid", fgColor="DCE6EC")  # フローの見出し行
C_GRP = PatternFill("solid", fgColor="F2F5F7")  # 群の帯
STATUS_FILL = {
    "PASS": PatternFill("solid", fgColor="D6EFD8"),
    "FAIL": PatternFill("solid", fgColor="F8D7DA"),
    "BLOCKED": PatternFill("solid", fgColor="FFF0CC"),
    "TODO": PatternFill("solid", fgColor="FFFFFF"),
    "SKIP": PatternFill("solid", fgColor="EDEDED"),
}
THIN = Side(style="thin", color="C9D3DA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WHITE_BOLD = Font(bold=True, color="FFFFFF")

#: 群の定義。ID の先頭 3 文字で振り分ける。
GROUPS: OrderedDict[str, tuple[str, str]] = OrderedDict(
    [
        ("J00", ("第0群 土台", "これが無いと他が始まらない。空の状態から最初の運営を作るまで")),
        ("J01", ("第0群 土台", "")),
        ("J02", ("第0群 土台", "")),
        ("J03", ("第0群 土台", "")),
        ("J10", ("第1群 入口", "新規登録から使い始めるまで。入力エラーも 1 つずつ踏む")),
        ("J11", ("第1群 入口", "")),
        ("J12", ("第1群 入口", "")),
        ("J13", ("第1群 入口", "")),
        ("J14", ("第1群 入口", "")),
        ("J15", ("第1群 入口", "")),
        ("J20", ("第2群 R-T08", "クライアント分離。**致命級** — ここが破れたら公開不可")),
        ("J21", ("第2群 R-T08", "")),
        ("J22", ("第2群 R-T08", "")),
        ("J23", ("第2群 R-T08", "")),
        ("J24", ("第2群 R-T08", "")),
        ("J30", ("第3群 案件を回す", "案件・工程・メンバー・承認・取り込み・議事録・ナレッジ")),
        ("J31", ("第3群 案件を回す", "")),
        ("J32", ("第3群 案件を回す", "")),
        ("J33", ("第3群 案件を回す", "")),
        ("J34", ("第3群 案件を回す", "")),
        ("J35", ("第3群 案件を回す", "")),
        ("J36", ("第3群 案件を回す", "")),
        ("J37", ("第3群 案件を回す", "")),
        ("J38", ("第3群 案件を回す", "")),
        ("J39", ("第3群 案件を回す", "")),
        ("J44", ("第3群 案件を回す", "")),
        ("J45", ("第3群 案件を回す", "")),
        ("J46", ("第3群 案件を回す", "")),
        ("J47", ("第3群 案件を回す", "")),
        ("J48", ("第3群 案件を回す", "")),
        ("J49", ("第3群 案件を回す", "")),
        ("J40", ("第4群 AI が仕事をする", "製品の中核。本人の PC・本人の Claude 契約で動く")),
        ("J41", ("第4群 AI が仕事をする", "")),
        ("J42", ("第4群 AI が仕事をする", "")),
        ("J43", ("第4群 AI が仕事をする", "")),
        ("J50", ("第5群 お金と出口", "申込・解約・退会")),
        ("J51", ("第5群 お金と出口", "")),
        ("J52", ("第5群 お金と出口", "")),
        ("J60", ("第6群 運営が見る", "健全性・混雑・監査ログ")),
        ("J61", ("第6群 運営が見る", "")),
        ("J62", ("第6群 運営が見る", "")),
        ("J63", ("第6群 運営が見る", "")),
        ("J64", ("第6群 運営が見る", "")),
    ]
)

#: フロー名（ID の先頭 3 文字 → 日本語）。一覧シートの主役。
FLOW_TITLES = {
    "J00": "空の状態から初回起動・最初の運営を作る",
    "J01": "運営が既定デザインテンプレを用意する",
    "J02": "運営が AI 社員テンプレとスキルを整える",
    "J03": "運営が法務文書を改訂し、現行版を切り替える",
    "J10": "新規登録 → 同意 → ワークスペース作成 → 着地",
    "J11": "サインアウト → 再サインイン → 状態が残っている",
    "J12": "規約が新版になった → 帯が出る → 同意する",
    "J13": "パスワードを忘れた → 再設定 → 旧セッションが死ぬ",
    "J14": "OAuth / Magic Link で入る → 拒否・二重使用が効く",
    "J15": "横断画面 (ウェルカム・通知・プロフィール・切替・検索) が実データを映す",
    "J20": "招待を発行する → クライアントが招待リンクで入る",
    "J21": "クライアントが自分の案件だけ見える",
    "J22": "クライアントが他人の案件へ越境できない",
    "J23": "クライアントがコメントし、社内に届く",
    "J24": "招待の期限切れ・失効が正しく効く",
    "J30": "案件を作る → 工程が用意される → 工程を進める",
    "J31": "メンバーを招待する → 権限どおりに使える",
    "J32": "承認待ちを承認する / 却下する",
    "J33": "既存案件を途中から取り込む",
    "J34": "議事録をアップロード → 文字起こし → 採用",
    "J35": "ナレッジを貯める → 検索で引ける → 昇格する",
    "J36": "ワークスペース設定 (名前・AI 学習・MCP トークン・削除) が効く",
    "J37": "AI 社員を編集する → 能力として見える → 運営改訂と共存する",
    "J38": "案件の設定・削除・30 日復元・アーカイブ",
    "J39": "案件シークレットを登録 → reveal → 削除 (平文を漏らさない)",
    "J44": "自動実行スケジュールを作る → 有効/無効 → 今すぐ実行 → 履歴",
    "J45": "モックを作る → 修正依頼 → 版管理 → コメント (凍結が守られる)",
    "J46": "成果物を見る → コメント → AI 修正提案 → 版 → 共有リンクで納品",
    "J47": "営業ドラフトを作る → PDF → dry_run/本送信 → 履歴",
    "J48": "タスクを作る → 再生 → スコア判定 → 承認/差戻 → 回収 (G-15)",
    "J49": "フェーズ提案 → 承認/却下 → タスク移動の影響解析 → 適用",
    "J40": "本人の PC を繋ぐ → 未接続時の案内が正しい",
    "J41": "チャットで相談 → 文脈が効く → 成果物になる",
    "J42": "承認モードで危険な操作が止まる",
    "J43": "混雑しても断られず順番待ちになる",
    "J50": "有料プランに申し込む",
    "J51": "解約する → 期間末まで使える",
    "J52": "退会する → 30 日猶予 → 復元できる",
    "J60": "運営が健全性・混雑実績を見て気づける",
    "J61": "運営が監査ログで「誰が何をしたか」を追える",
    "J62": "運営が事業 KPI (目標・チャネル・コスト) を記録して見る",
    "J63": "運営がユーザーを横断して見てサポート連絡を送る",
    "J64": "運営デフォルト・ナレッジとキュレーション (匿名化) を回す",
}

ROLE_JA = {
    "admin": "運営",
    "owner": "オーナー",
    "member": "メンバー",
    "client_portal": "クライアント",
    "guest": "未サインイン",
}
BRANCH_JA = {
    "happy": "正常系",
    "validation": "入力エラー",
    "permission": "権限",
    "empty": "0件・未設定",
    "limit": "上限・期限",
    "conflict": "競合・凍結",
    "cancel": "取消・拒否",
    "isolation": "越境の遮断",
}


def flow_key(row_id: str) -> str:
    return row_id.split("-")[0]


def place_of(note: str) -> str:
    m = re.search(r"場所=(\w+)", note or "")
    return {"cloud": "この環境", "mac": "Mac (AI 実行)"}.get(m.group(1) if m else "", "")


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = C_HEAD
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_overview(wb: Workbook, rows: list[dict]) -> None:
    """最初に開くシート。**全フロー (43 本) を 1 行ずつ**。ここだけ見れば全体が分かる。"""
    ws = wb.create_sheet("フロー一覧", 0)
    ws.append(
        [
            "群",
            "ID",
            "フロー",
            "主なロール",
            "分岐",
            "ステップ",
            "済",
            "進捗",
            "状態",
            "実行場所",
            "依存",
        ]
    )
    style_header(ws, 1, 11)
    set_widths(ws, [18, 7, 46, 13, 22, 8, 6, 10, 12, 14, 12])
    ws.freeze_panes = "A2"

    seen: OrderedDict[str, list[dict]] = OrderedDict()
    for r in rows:
        seen.setdefault(flow_key(r["id"]), []).append(r)

    last_group = None
    for key, items in seen.items():
        group = GROUPS.get(key, ("その他", ""))[0]
        if group != last_group:
            ws.append([group] + [""] * 10)
            for c in range(1, 12):
                ws.cell(row=ws.max_row, column=c).fill = C_GRP
                ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
            last_group = group
        done = sum(1 for i in items if i["status"] == "PASS")
        n = len(items)
        roles = sorted({ROLE_JA.get(i["role"], i["role"]) for i in items})
        branches = sorted({BRANCH_JA.get(i["branch"], i["branch"]) for i in items})
        deps = sorted({d.split("-")[0] for i in items for d in (i.get("depends_on") or []) if d})
        state = "完了" if done == n else ("着手中" if done else "未着手")
        ws.append(
            [
                "",
                key,
                FLOW_TITLES.get(key, ""),
                " / ".join(roles),
                " / ".join(branches),
                n,
                done,
                f"{done}/{n}",
                state,
                place_of(items[0]["note"]),
                ", ".join(d for d in deps if d != key) or "—",
            ]
        )
        row = ws.max_row
        ws.cell(row=row, column=9).fill = STATUS_FILL["PASS" if done == n else "TODO"]
        for c in range(1, 12):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)


def build_group_sheets(wb: Workbook, rows: list[dict]) -> None:
    """群ごとのタブ。フローの見出し行 → その配下にステップ、で塊が見える。"""
    by_group: OrderedDict[str, list[dict]] = OrderedDict()
    for r in rows:
        g = GROUPS.get(flow_key(r["id"]), ("その他", ""))[0]
        by_group.setdefault(g, []).append(r)

    for group, items in by_group.items():
        ws = wb.create_sheet(group[:31])
        ws.append(
            [
                "ID",
                "ロール",
                "分岐",
                "やること",
                "操作手順",
                "これが出れば成功",
                "状態",
                "証拠",
                "備考",
            ]
        )
        style_header(ws, 1, 9)
        set_widths(ws, [9, 12, 12, 34, 40, 46, 10, 22, 26])
        ws.freeze_panes = "A2"

        cur = None
        for r in sorted(items, key=lambda x: x["order"]):
            key = flow_key(r["id"])
            if key != cur:
                ws.append([key, FLOW_TITLES.get(key, ""), "", "", "", "", "", "", ""])
                hr = ws.max_row
                ws.merge_cells(start_row=hr, start_column=2, end_row=hr, end_column=9)
                for c in range(1, 10):
                    ws.cell(row=hr, column=c).fill = C_FLOW
                    ws.cell(row=hr, column=c).font = Font(bold=True)
                    ws.cell(row=hr, column=c).border = BORDER
                cur = key
            ws.append(
                [
                    r["id"],
                    ROLE_JA.get(r["role"], r["role"]),
                    BRANCH_JA.get(r["branch"], r["branch"]),
                    r["action"],
                    r["steps"],
                    r["expected"],
                    r["status"],
                    r["evidence"],
                    r["note"],
                ]
            )
            row = ws.max_row
            ws.cell(row=row, column=7).fill = STATUS_FILL.get(r["status"], STATUS_FILL["TODO"])
            for c in range(1, 10):
                ws.cell(row=row, column=c).border = BORDER
                ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)


def build_plan(wb: Workbook, rows: list[dict]) -> None:
    """機械可読シート。journey_workbook.py の update/status がここを読む。列名は変えない。"""
    ws = wb.create_sheet("Plan")
    cols = [
        "id",
        "order",
        "role",
        "phase",
        "data_condition",
        "branch",
        "depends_on",
        "action",
        "steps",
        "expected",
        "status",
        "evidence",
        "note",
    ]
    ws.append(cols)
    style_header(ws, 1, len(cols))
    set_widths(ws, [10, 7, 13, 11, 14, 11, 14, 30, 34, 38, 10, 20, 24])
    ws.freeze_panes = "A2"
    for r in sorted(rows, key=lambda x: x["order"]):
        ws.append(
            [
                r["id"],
                r["order"],
                r["role"],
                r["phase"],
                r["data_condition"],
                r["branch"],
                ",".join(r.get("depends_on") or []),
                r["action"],
                r["steps"],
                r["expected"],
                r["status"],
                r["evidence"],
                r["note"],
            ]
        )


def build_roles(wb: Workbook, plan: dict) -> None:
    """Roles シート。中身は plan.json の discovered.role_meta (正本) から取る — ここに手で書かない (GAP-248)。"""
    ws = wb.create_sheet("Roles")
    ws.append(["ロール", "入口", "ゴール", "供給するもの", "消費するもの"])
    style_header(ws, 1, 5)
    set_widths(ws, [22, 40, 36, 40, 40])
    discovered = plan.get("discovered", {}) or {}
    role_meta = discovered.get("role_meta", {}) or {}
    for role in discovered.get("roles", []):
        meta = role_meta.get(role, {}) or {}
        ws.append(
            [
                f"{ROLE_JA.get(role, role)} ({role})",
                meta.get("how_to_enter", ""),
                meta.get("goal", ""),
                meta.get("provides", ""),
                meta.get("consumes", ""),
            ]
        )
        for c in range(1, 6):
            ws.cell(row=ws.max_row, column=c).border = BORDER
            ws.cell(row=ws.max_row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
    ws.append([])
    ws.append(["実行順の原則", discovered.get("dependency_note", "")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)


def build_summary(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Summary")
    ws.append(["項目", "値", "割合"])
    style_header(ws, 1, 3)
    set_widths(ws, [26, 14, 10])
    n = len(rows)
    counts = {
        s: sum(1 for r in rows if r["status"] == s)
        for s in ("PASS", "FAIL", "BLOCKED", "TODO", "SKIP")
    }
    ws.append(["総ステップ", n, ""])
    for s, v in counts.items():
        ws.append([s, v, f"{round(v / n * 100)}%" if n else "0%"])
        ws.cell(row=ws.max_row, column=1).fill = STATUS_FILL[s]
    done = counts["PASS"] == n
    ws.append(["DONE?", "YES" if done else "NO", ""])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True, color="1B7F3B" if done else "B3261E")
    ws.append([])
    ws.append(["ロール別", "ステップ", "済"])
    style_header(ws, ws.max_row, 3)
    for role in ROLE_JA:
        rs = [r for r in rows if r["role"] == role]
        if rs:
            ws.append([ROLE_JA[role], len(rs), sum(1 for r in rs if r["status"] == "PASS")])
    ws.append([])
    ws.append(["分岐別", "ステップ", "済"])
    style_header(ws, ws.max_row, 3)
    for br in BRANCH_JA:
        rs = [r for r in rows if r["branch"] == br]
        if rs:
            ws.append([BRANCH_JA[br], len(rs), sum(1 for r in rs if r["status"] == "PASS")])
    ws.append([])
    ws.append(["実行場所", "ステップ", "済"])
    style_header(ws, ws.max_row, 3)
    for label in ("この環境", "Mac (AI 実行)"):
        rs = [r for r in rows if place_of(r["note"]) == label]
        if rs:
            ws.append([label, len(rs), sum(1 for r in rs if r["status"] == "PASS")])


def carry_status(rows: list[dict], out: Path) -> int:
    """既存 Excel の Plan シートから status/evidence を引き継ぐ（実行結果を捨てない）。"""
    if not out.exists():
        return 0
    wb = load_workbook(out)
    if "Plan" not in wb.sheetnames:
        return 0
    ws = wb["Plan"]
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return 0
    head = list(data[0])
    got = {}
    for r in data[1:]:
        if r and r[head.index("id")]:
            got[r[head.index("id")]] = (r[head.index("status")], r[head.index("evidence")])
    n = 0
    for r in rows:
        if r["id"] in got:
            st, ev = got[r["id"]]
            if st and st != "TODO":
                r["status"], r["evidence"], n = st, ev or "", n + 1
    return n


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--plan", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--carry-status", action="store_true", help="既存 Excel の実行結果を引き継ぐ")
    args = ap.parse_args()

    plan = json.loads(Path(args.plan).read_text(encoding="utf-8"))
    rows = plan["rows"]
    out = Path(args.out)
    carried = carry_status(rows, out) if args.carry_status else 0

    wb = Workbook()
    wb.remove(wb.active)
    build_overview(wb, rows)
    build_group_sheets(wb, rows)
    build_plan(wb, rows)
    build_roles(wb, plan)
    build_summary(wb, rows)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(
        f"✓ {out}  ({len(rows)} ステップ / シート {len(wb.sheetnames)}: {', '.join(wb.sheetnames)})"
    )
    if carried:
        print(f"  既存の実行結果 {carried} 件を引き継ぎました")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
